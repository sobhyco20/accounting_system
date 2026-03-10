from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.http import JsonResponse
from .models import Customer, SalesInvoice, SalesInvoiceItem, SalesReturn, CustomerPayment, CustomerGroup
from django.db.models import Sum, F, ExpressionWrapper, DecimalField, Q
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.http import HttpResponse
import pandas as pd


@property
def total_with_tax(self):
    return self.total_after_tax  # نفس القيمة


def sales_invoice_list(request):
    invoices = SalesInvoice.objects.select_related('customer').all()
    return render(request, 'sales/invoice_list.html', {'invoices': invoices})


def sales_invoice_detail(request, pk):
    invoice = get_object_or_404(SalesInvoice, pk=pk)
    items = invoice.items.all()
    return render(request, 'sales/invoice_detail.html', {'invoice': invoice, 'items': items})


def sales_return_list(request):
    returns = SalesReturn.objects.select_related('customer', 'original_invoice').all()
    return render(request, 'sales/return_list.html', {'returns': returns})


def customer_list(request):
    customers = Customer.objects.select_related('group').all()
    return render(request, 'sales/customers.html', {'customers': customers})


def customer_detail(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    invoices = SalesInvoice.objects.filter(customer=customer)
    returns = SalesReturn.objects.filter(customer=customer)
    payments = CustomerPayment.objects.filter(customer=customer)

    context = {
        'customer': customer,
        'invoices': invoices,
        'returns': returns,
        'payments': payments,
    }
    return render(request, 'sales/customer_detail.html', context)

#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
from django.shortcuts import render
from django.utils import timezone
from django.db.models import Sum
from .models import Customer, SalesInvoice, CustomerPayment
from datetime import datetime, date

def aging_report(request):
    to_date_str = request.GET.get('to_date')
    to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date() if to_date_str else timezone.now().date()
    from_date = date(2025, 1, 1)  # بداية السنة المالية

    aging_data = []
    customers = Customer.objects.all()

    for customer in customers:
        ranges = {
            'range_0_30': 0,
            'range_31_60': 0,
            'range_61_90': 0,
            'range_91_120': 0,
            'range_121_plus': 0,
        }
        total_due = 0
        opening_balance = customer.opening_debit - customer.opening_credit

        if opening_balance > 0:
            days = (to_date - from_date).days
            total_due += opening_balance
            if days <= 30:
                ranges['range_0_30'] += opening_balance
            elif days <= 60:
                ranges['range_31_60'] += opening_balance
            elif days <= 90:
                ranges['range_61_90'] += opening_balance
            elif days <= 120:
                ranges['range_91_120'] += opening_balance
            else:
                ranges['range_121_plus'] += opening_balance

        # فواتير المبيعات
        invoices = SalesInvoice.objects.filter(customer=customer, date__lte=to_date)
        for invoice in invoices:
            days = (to_date - invoice.date).days
            paid = CustomerPayment.objects.filter(
                invoice=invoice, date__lte=to_date
            ).aggregate(total=Sum('amount'))['total'] or 0

            remaining = invoice.total_with_tax_value - paid

            if remaining > 0:
                total_due += remaining
                if days <= 30:
                    ranges['range_0_30'] += remaining
                elif days <= 60:
                    ranges['range_31_60'] += remaining
                elif days <= 90:
                    ranges['range_61_90'] += remaining
                elif days <= 120:
                    ranges['range_91_120'] += remaining
                else:
                    ranges['range_121_plus'] += remaining

        if total_due > 0:
            aging_data.append({
                'customer': customer,
                'opening_balance': opening_balance,
                'total': total_due,
                **ranges,
            })

    return render(request, 'sales/reports/aging_report.html', {
        'aging_data': aging_data,
        'to_date': to_date,
    })


#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

from django.shortcuts import render, get_object_or_404
from datetime import datetime
from .models import Customer, SalesInvoice, SalesReturn, CustomerPayment
from decimal import Decimal

def customer_ledger(request):
    customers = Customer.objects.all()
    customer_id = request.GET.get('customer')
    from_date = request.GET.get('from')
    to_date = request.GET.get('to')

    entries = []
    total_debit = Decimal("0.00")
    total_credit = Decimal("0.00")
    final_balance = Decimal("0.00")
    customer = None

    if customer_id:
        customer = get_object_or_404(Customer, id=customer_id)

        # تحويل التواريخ إلى datetime أو استخدام نطاق واسع إذا لم تُحدد
        from_date_obj = datetime.strptime(from_date, "%Y-%m-%d") if from_date else datetime(2000, 1, 1)
        to_date_obj = datetime.strptime(to_date, "%Y-%m-%d") if to_date else datetime(2100, 1, 1)

        # 🟡 رصيد افتتاحي (بدون أقواس لأنهم حقول وليست دوال)
        opening_debit = customer.opening_debit
        opening_credit = customer.opening_credit
        opening_balance = opening_debit - opening_credit

        entries.append({
            'date': from_date_obj.date(),
            'description': 'رصيد افتتاحي',
            'debit': opening_debit,
            'credit': opening_credit,
            'balance': opening_balance
        })

        # 🟢 فواتير مبيعات
        invoices = SalesInvoice.objects.filter(customer=customer, date__range=[from_date_obj, to_date_obj])
        for inv in invoices:
            entries.append({
                'date': inv.date,
                'description': f"فاتورة مبيعات رقم {inv.number}",
                'debit': inv.total_with_tax(),
                'credit': Decimal("0.00"),
            })

        # 🔵 مردودات مبيعات
        returns = SalesReturn.objects.filter(customer=customer, date__range=[from_date_obj, to_date_obj])
        for ret in returns:
            entries.append({
                'date': ret.date,
                'description': f"مردود مبيعات رقم {ret.number}",
                'debit': Decimal("0.00"),
                'credit': ret.total_after_tax(),  # استدعاء الدالة بين قوسين
            })

        # 🔴 سداد عملاء
        payments = CustomerPayment.objects.filter(customer=customer, date__range=[from_date_obj, to_date_obj])
        for pay in payments:
            entries.append({
                'date': pay.date,
                'description': f"سداد رقم {pay.id}",
                'debit': Decimal("0.00"),
                'credit': pay.amount,
            })

        # ترتيب الحركات بالتاريخ ثم إعادة احتساب الرصيد
        entries.sort(key=lambda x: x['date'])
        balance = opening_balance
        for entry in entries[1:]:  # نبدأ بعد الرصيد الافتتاحي
            balance += entry['debit'] - entry['credit']
            entry['balance'] = balance
            total_debit += entry['debit']
            total_credit += entry['credit']
        final_balance = balance

    return render(request, 'sales/reports/customer_statement.html', {
        'customers': customers,
        'customer': customer,
        'entries': entries,
        'from_date': from_date,
        'to_date': to_date,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'final_balance': final_balance,
    })

#-----------------------------------------------------------------------------
from django.shortcuts import render
from django.utils.dateparse import parse_date
from sales.models import SalesInvoice, SalesReturn

def sales_report_view(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    sales_invoices = SalesInvoice.objects.all()
    sales_returns = SalesReturn.objects.all()

    if from_date and to_date:
        from_date_parsed = parse_date(from_date)
        to_date_parsed = parse_date(to_date)
        sales_invoices = sales_invoices.filter(date__range=(from_date_parsed, to_date_parsed))
        sales_returns = sales_returns.filter(date__range=(from_date_parsed, to_date_parsed))

    total_sales = sum(inv.total_with_tax for inv in sales_invoices)
    total_returns = sum(ret.total_with_tax for ret in sales_returns)

    context = {
        'sales_invoices': sales_invoices,
        'sales_returns': sales_returns,
        'total_sales': total_sales,
        'total_returns': total_returns,
    }
    return render(request, 'sales/reports/sales_report.html', context)




#------------------------------------------------------------------------------------------------------
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render

@staff_member_required
def reports_home(request):
    return render(request, 'sales/reports/reports_home.html')


###########################################################################################################
def sales_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    export = request.GET.get('export')

    invoices = SalesInvoice.objects.select_related('customer').prefetch_related('items').all()

    if start_date:
        invoices = invoices.filter(date__gte=start_date)
    if end_date:
        invoices = invoices.filter(date__lte=end_date)

    # قائمة الفواتير + الإجمالي
    invoice_data = []
    total_amount = 0

    for inv in invoices:
        amount = sum(
            (item.quantity * item.unit_price) + item.tax_amount
            for item in inv.items.all()
        )
        invoice_data.append({
            'number': inv.number,
            'date': inv.date,
            'customer': inv.customer.name,
            'amount': amount,
        })
        total_amount += amount

    # Excel export
    if export:
        df = pd.DataFrame(invoice_data)
        df.columns = ['رقم الفاتورة', 'تاريخ الفاتورة', 'اسم العميل', 'إجمالي الفاتورة']
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=sales_report.xlsx'
        df.to_excel(response, index=False)
        return response

    context = {
        'invoices': invoice_data,
        'total_amount': total_amount,
        'from_date': start_date,
        'to_date': end_date,
    }
    return render(request, 'sales/reports/sales_report.html', context)

##################################################################################################################
from django.shortcuts import render
from datetime import datetime
from .models import Customer, SalesInvoice, SalesReturn

def customer_balances_report(request):
    from_date = request.GET.get('from_date') or '2026-01-01'
    to_date = request.GET.get('to_date') or datetime.today().strftime('%Y-%m-%d')

    customers = Customer.objects.all()
    report_data = []
    totals = {
        'opening_debit': 0,
        'opening_credit': 0,
        'debit_movements': 0,
        'credit_movements': 0,
        'end_debit': 0,
        'end_credit': 0,
    }

    for customer in customers:
        opening_debit = customer.opening_debit or 0
        opening_credit = customer.opening_credit or 0

        invoices = SalesInvoice.objects.filter(
            customer=customer, date__range=[from_date, to_date]
        )

        returns = SalesReturn.objects.filter(
            customer=customer, date__range=[from_date, to_date]
        )

        # جمع القيم باستدعاء الدالة () إذا لم تكن خاصية
        debit_movements = sum(inv.total_after_tax() for inv in invoices)
        credit_movements = sum(ret.total_after_tax() for ret in returns)

        end_debit = opening_debit + debit_movements
        end_credit = opening_credit + credit_movements
        final_balance = end_debit - end_credit
        status = 'مدين' if final_balance > 0 else 'دائن' if final_balance < 0 else 'متزن'

        report_data.append({
            'code': customer.code,
            'name': customer.name,
            'opening_debit': opening_debit,
            'opening_credit': opening_credit,
            'debit_movements': debit_movements,
            'credit_movements': credit_movements,
            'end_debit': end_debit,
            'end_credit': end_credit,
            'final_balance': abs(final_balance),
            'status': status,
            'invoices': [
                {
                    'number': inv.number,
                    'date': inv.date,
                    'amount': inv.total_after_tax(),  # يجب استدعاء الدالة
                } for inv in invoices
            ]
        })

        totals['opening_debit'] += opening_debit
        totals['opening_credit'] += opening_credit
        totals['debit_movements'] += debit_movements
        totals['credit_movements'] += credit_movements
        totals['end_debit'] += end_debit
        totals['end_credit'] += end_credit

    context = {
        'customers': report_data,
        'from_date': from_date,
        'to_date': to_date,
        'totals': totals,
    }

    return render(request, 'sales/reports/customer_balances.html', context)

#-------------------------------------------------------------------------------------

# سداد العملاء

from django.http import JsonResponse
from django.views import View
from .models import SalesInvoice, CustomerPayment

from django.db.models import Sum, F

class LoadInvoicesView(View):
    def get(self, request, *args, **kwargs):
        customer_id = request.GET.get('customer_id')
        invoices = SalesInvoice.objects.filter(customer_id=customer_id)

        # قائمة الفواتير مع السداد الحالي لها
        data = []
        for invoice in invoices:
            total_paid = CustomerPayment.objects.filter(invoice=invoice).aggregate(total=Sum('amount'))['total'] or 0
            remaining = invoice.total_with_tax_value - total_paid

            if remaining > 0:  # فقط الفواتير التي بها متبقي
                data.append({
                    'id': invoice.id,
                    'number': invoice.number,
                    'total': remaining
                })

        return JsonResponse({'invoices': data})
########################################################################################################################
from django.shortcuts import render
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO

from sales.models import Customer, SalesInvoice, SalesInvoiceItem, SalesReturn, SalesReturnItem, CustomerPayment
from inventory.models import Product

# ------------------------ تقرير المبيعات حسب العميل ------------------------
from django.db.models import Sum, Count
from django.utils.dateparse import parse_date
from django.http import HttpResponse
from openpyxl import Workbook

def sales_by_customer_report_view(request):
    customers = Customer.objects.all()
    customer_id = request.GET.get("customer")
    date_from = request.GET.get("from_date") or "2026-01-01"
    date_to = request.GET.get("to_date") or "2026-12-31"

    # فلترة حسب التاريخ والعميل
    sales_qs = SalesInvoice.objects.filter(date__range=(date_from, date_to))
    returns_qs = SalesReturn.objects.filter(date__range=(date_from, date_to))

    if customer_id:
        sales_qs = sales_qs.filter(customer_id=customer_id)
        returns_qs = returns_qs.filter(customer_id=customer_id)

    # تجميع البيانات لكل عميل
    report_data = []
    customers_filtered = customers.filter(id=customer_id) if customer_id else customers

    total_net = 0
    for cust in customers_filtered:
        cust_sales = sales_qs.filter(customer=cust)
        cust_returns = returns_qs.filter(customer=cust)

        sales_total = cust_sales.aggregate(total=Sum('total_with_tax_value'))['total'] or 0
        returns_total = cust_returns.aggregate(total=Sum('total_with_tax_value'))['total'] or 0
        invoice_count = cust_sales.count()
        return_count = cust_returns.count()
        net_total = sales_total - returns_total

        report_data.append({
            'customer': cust.name,
            'invoice_count': invoice_count,
            'sales_total': sales_total,
            'return_count': return_count,
            'returns_total': returns_total,
            'net_total': net_total,
        })

        total_net += net_total

    # تصدير Excel
    if request.GET.get('export') == '1':
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير المبيعات حسب العميل"

        ws.append(["العميل", "عدد الفواتير", "إجمالي المبيعات", "عدد المرتجعات", "إجمالي المرتجعات", "الصافي"])

        for row in report_data:
            ws.append([
                row['customer'],
                row['invoice_count'],
                float(row['sales_total']),
                row['return_count'],
                float(row['returns_total']),
                float(row['net_total']),
            ])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename=sales_by_customer.xlsx'
        wb.save(response)
        return response

    return render(request, "sales/reports/sales_by_customer_report.html", {
        'report_data': report_data,
        'total_amount': total_net,
        'customers': customers,
        'selected_customer': int(customer_id) if customer_id else None,
        'from_date': date_from,
        'to_date': date_to,
    })


# ------------------------ كشف حساب عميل ------------------------
def customer_ledger_report_view(request):
    customers = Customer.objects.all()
    customer_id = request.GET.get("customer")
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")
    export = request.GET.get("export")

    ledger = []
    opening_balance = 0

    if customer_id and date_from and date_to:
        customer = Customer.objects.get(id=customer_id)
        opening_balance = customer.opening_debit - customer.opening_credit

        sales_before = SalesInvoice.objects.filter(customer=customer, date__lt=date_from)
        returns_before = SalesReturn.objects.filter(customer=customer, date__lt=date_from)
        payments_before = CustomerPayment.objects.filter(customer=customer, date__lt=date_from)

        total_sales_before = sum(
            sum(item.total_before_tax + item.tax_amount for item in inv.items.all())
            for inv in sales_before
        )
        total_returns_before = sum(
            sum(item.total_before_tax + item.tax_amount for item in ret.details.all())
            for ret in returns_before
        )
        total_payments_before = sum(pay.total_amount for pay in payments_before)

        opening_balance += total_sales_before - total_returns_before - total_payments_before

        movements = []

        sales = SalesInvoice.objects.filter(customer=customer, date__gte=date_from, date__lte=date_to)
        for inv in sales:
            total_invoice = sum(item.total_before_tax + item.tax_amount for item in inv.items.all())
            movements.append({
                'date': inv.date,
                'description': f"فاتورة مبيعات رقم {inv.number}",
                'debit': total_invoice,
                'credit': 0
            })

        returns = SalesReturn.objects.filter(customer=customer, date__gte=date_from, date__lte=date_to)
        for ret in returns:
            total_return = sum(item.total_before_tax + item.tax_amount for item in ret.details.all())
            movements.append({
                'date': ret.date,
                'description': f"مرتجع مبيعات رقم {ret.number}",
                'debit': 0,
                'credit': total_return
            })

        payments = CustomerPayment.objects.filter(customer=customer, date__gte=date_from, date__lte=date_to)
        for pay in payments:
            movements.append({
                'date': pay.date,
                'description': f"سداد رقم {pay.number}",
                'debit': 0,
                'credit': pay.amount
            })

        movements = sorted(movements, key=lambda x: x['date'])

        balance = opening_balance
        ledger.append({
            'date': "رصيد أول المدة",
            'description': '',
            'debit': opening_balance if opening_balance > 0 else 0,
            'credit': abs(opening_balance) if opening_balance < 0 else 0,
            'balance': balance
        })

        for move in movements:
            balance += move['debit'] - move['credit']
            move['balance'] = balance
            ledger.append(move)

        if export == 'excel':
            return export_customer_ledger_to_excel(ledger, customer, opening_balance)

    context = {
        'title': "كشف حساب عميل",
        'customers': customers,
        'ledger': ledger,
        'opening_balance': opening_balance,
        'selected_customer': int(customer_id) if customer_id else None,
        'date_from': date_from,
        'date_to': date_to
    }
    return render(request, "sales/reports/customer_ledger_report.html", context)

def export_customer_ledger_to_excel(ledger, customer, opening_balance):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"كشف حساب"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal="center")
    bold_font = Font(bold=True)

    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = f"كشف حساب العميل: {customer.name}"
    cell.font = Font(bold=True, size=14)
    cell.alignment = center_alignment

    ws.merge_cells('A2:E2')
    cell = ws['A2']
    cell.value = f"رصيد أول المدة: {opening_balance:,.2f}"
    cell.font = Font(bold=True, size=12)
    cell.alignment = center_alignment

    headers = ["التاريخ", "البيان", "مدين", "دائن", "الرصيد التراكمي"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    for row_num, row in enumerate(ledger, 5):
        ws.cell(row=row_num, column=1, value=str(row['date'])).alignment = center_alignment
        ws.cell(row=row_num, column=2, value=row['description']).alignment = Alignment(horizontal="right")
        ws.cell(row=row_num, column=3, value=row['debit']).number_format = '#,##0.00'
        ws.cell(row=row_num, column=4, value=row['credit']).number_format = '#,##0.00'
        ws.cell(row=row_num, column=5, value=row['balance']).number_format = '#,##0.00'
        for col in range(1, 6):
            ws.cell(row=row_num, column=col).border = thin_border

    current_balance = ledger[-1]['balance'] if ledger else opening_balance
    summary_row = len(ledger) + 6
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=4)
    ws.cell(row=summary_row, column=1, value="الرصيد الحالي").font = bold_font
    ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=summary_row, column=5, value=current_balance).font = bold_font
    ws.cell(row=summary_row, column=5).number_format = '#,##0.00'
    ws.cell(row=summary_row, column=5).alignment = center_alignment

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length + 5

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename=ledger_{customer.name}.xlsx'
    buffer = BytesIO()
    wb.save(buffer)
    response.write(buffer.getvalue())
    return response

# ---------------------------------------------
# تقرير المبيعات حسب الأصناف والعملاء
# ---------------------------------------------
def sales_by_product_and_customer_report_view(request):
    products = Product.objects.all()
    customers = Customer.objects.all()

    date_from = request.GET.get("date_from", "2026-01-01")
    date_to = request.GET.get("date_to", "2026-12-31")
    product_id = request.GET.get("product")
    customer_id = request.GET.get("customer")
    export = request.GET.get("export")

    sales_items = SalesInvoiceItem.objects.select_related('invoice', 'product').filter(
        invoice__date__range=[date_from, date_to]
    )
    if product_id:
        sales_items = sales_items.filter(product_id=product_id)
    if customer_id:
        sales_items = sales_items.filter(invoice__customer_id=customer_id)

    returns_items = SalesReturnItem.objects.select_related('sales_return', 'product').filter(
        sales_return__date__range=[date_from, date_to]
    )
    if product_id:
        returns_items = returns_items.filter(product_id=product_id)
    if customer_id:
        returns_items = returns_items.filter(sales_return__customer_id=customer_id)

    report_dict = {}

    for item in sales_items:
        key = (item.product.id, item.invoice.customer.id)
        report_dict.setdefault(key, {
            'product': item.product.name,
            'customer': item.invoice.customer.name,
            'quantity': 0, 'total_before_tax': 0, 'tax_amount': 0, 'total_with_tax': 0
        })
        report_dict[key]['quantity'] += item.quantity
        report_dict[key]['total_before_tax'] += item.total_before_tax
        report_dict[key]['tax_amount'] += item.tax_amount
        report_dict[key]['total_with_tax'] += item.total_before_tax + item.tax_amount

    for item in returns_items:
        key = (item.product.id, item.sales_return.customer.id)
        report_dict.setdefault(key, {
            'product': item.product.name,
            'customer': item.sales_return.customer.name,
            'quantity': 0, 'total_before_tax': 0, 'tax_amount': 0, 'total_with_tax': 0
        })
        report_dict[key]['quantity'] -= item.quantity
        report_dict[key]['total_before_tax'] -= item.total_before_tax
        report_dict[key]['tax_amount'] -= item.tax_amount
        report_dict[key]['total_with_tax'] -= item.total_before_tax + item.tax_amount

    report = list(report_dict.values())

    if export == "excel":
        return export_sales_by_product_and_customer_excel(report, date_from, date_to)

    context = {
        'products': products,
        'customers': customers,
        'report': report,
        'date_from': date_from,
        'date_to': date_to
    }
    return render(request, "sales/reports/sales_by_product_and_customer.html", context)


def export_sales_by_product_and_customer_excel(report, date_from, date_to):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "مبيعات الأصناف"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal="center")

    ws.merge_cells('A1:F1')
    ws['A1'].value = f"تقرير المبيعات من {date_from} إلى {date_to}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_alignment

    headers = ["الصنف", "العميل", "الكمية", "الإجمالي قبل الضريبة", "الضريبة", "الإجمالي بعد الضريبة"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    for row_num, row in enumerate(report, 4):
        ws.cell(row=row_num, column=1, value=row['product'])
        ws.cell(row=row_num, column=2, value=row['customer'])
        ws.cell(row=row_num, column=3, value=row['quantity'])
        ws.cell(row=row_num, column=4, value=row['total_before_tax']).number_format = '#,##0.00'
        ws.cell(row=row_num, column=5, value=row['tax_amount']).number_format = '#,##0.00'
        ws.cell(row=row_num, column=6, value=row['total_with_tax']).number_format = '#,##0.00'
        for col in range(1, 7):
            ws.cell(row=row_num, column=col).border = thin_border

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length + 5

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename=sales_by_product_and_customer.xlsx'
    buffer = BytesIO()
    wb.save(buffer)
    response.write(buffer.getvalue())
    return response


# ---------------------------------------------
# تقرير أعمار الديون
# ---------------------------------------------
from datetime import datetime
from io import BytesIO
from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from sales.models import Customer, SalesInvoice, SalesReturn, CustomerPayment


def aging_report_view(request):
    customers = Customer.objects.all()
    customer_id = request.GET.get("customer")
    date_from = request.GET.get("date_from", "2026-01-01")
    date_to = request.GET.get("date_to", "2026-12-31")
    export = request.GET.get("export")

    if customer_id:
        customers = customers.filter(id=customer_id)

    report = []
    for customer in customers:
        opening_balance = customer.opening_debit - customer.opening_credit

        invoices = SalesInvoice.objects.filter(customer=customer, date__range=[date_from, date_to])
        returns = SalesReturn.objects.filter(customer=customer, date__range=[date_from, date_to])
        payments = CustomerPayment.objects.filter(customer=customer, date__range=[date_from, date_to])

        total_invoices = sum(
            sum(item.total_before_tax + item.tax_amount for item in inv.items.all()) for inv in invoices
        )
        total_returns = sum(
            sum(item.total_before_tax + item.tax_amount for item in ret.details.all()) for ret in returns
        )
        total_payments = sum(p.amount for p in payments)

        total_balance = opening_balance + total_invoices - total_returns - total_payments

        today = datetime.strptime(date_to, "%Y-%m-%d").date()
        aging = {'0-30': 0, '31-60': 0, '61-90': 0, '91-120': 0, '120+': 0}

        for inv in invoices:
            invoice_date = inv.date
            total_inv = sum(item.total_before_tax + item.tax_amount for item in inv.items.all())
            days = (today - invoice_date).days

            if days <= 30:
                aging['0-30'] += total_inv
            elif days <= 60:
                aging['31-60'] += total_inv
            elif days <= 90:
                aging['61-90'] += total_inv
            elif days <= 120:
                aging['91-120'] += total_inv
            else:
                aging['120+'] += total_inv

        report.append({
            'customer': customer.name,
            'total': total_balance,
            **aging
        })

    if export == "excel":
        return export_aging_report_to_excel(report, date_from, date_to)

    context = {
        'customers': Customer.objects.all(),
        'report': report,
        'date_from': date_from,
        'date_to': date_to
    }
    return render(request, "reports/aging_report.html", context)


def export_aging_report_to_excel(report, date_from, date_to):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "أعمار الديون"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal="center")

    ws.merge_cells('A1:H1')
    ws['A1'].value = f"تقرير أعمار الديون من {date_from} إلى {date_to}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_alignment

    headers = ["العميل", "الرصيد", "0-30", "31-60", "61-90", "91-120", "120+"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    for row_num, row in enumerate(report, 4):
        ws.cell(row=row_num, column=1, value=row['customer'])
        ws.cell(row=row_num, column=2, value=row['total']).number_format = '#,##0.00'
        ws.cell(row=row_num, column=3, value=row['0-30']).number_format = '#,##0.00'
        ws.cell(row=row_num, column=4, value=row['31-60']).number_format = '#,##0.00'
        ws.cell(row=row_num, column=5, value=row['61-90']).number_format = '#,##0.00'
        ws.cell(row=row_num, column=6, value=row['91-120']).number_format = '#,##0.00'
        ws.cell(row=row_num, column=7, value=row['120+']).number_format = '#,##0.00'

        for col in range(1, 8):
            ws.cell(row=row_num, column=col).border = thin_border

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length + 5

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename=aging_report.xlsx'
    buffer = BytesIO()
    wb.save(buffer)
    response.write(buffer.getvalue())
    return response





# sales/views.py

from django.template.loader import render_to_string
from django.templatetags.static import static
from weasyprint import HTML
from django.http import HttpResponse
from .models import SalesInvoice
from core.utils import get_company
from .utils.zatca_qr import generate_invoice_qr_base64


def invoice_pdf_view(request, invoice_id):
    invoice = SalesInvoice.objects.get(id=invoice_id)
    company = get_company()

    font_url = request.build_absolute_uri(static('fonts/Amiri-Regular.ttf'))

    # طباعة القيم للمساعدة في التحقق
    print("🔍 Generating QR for invoice:")
    print("Seller:", company.name)
    print("VAT:", company.tax_number)
    print("Date:", invoice.date)
    print("Total with VAT:", invoice.total_with_tax_value)
    print("VAT Total:", invoice.total_tax_value)

                
    qr_code_base64 = generate_invoice_qr_base64(
        seller_name=company.name,
        vat_number=company.tax_number,
        timestamp=str(invoice.date),
        invoice_total=str(invoice.total_with_tax_value),
        vat_total=str(invoice.total_tax_value)
    )



    html_string = render_to_string('sales/invoice_pdf.html', {
        'invoice': invoice,
        'company': company,
        'font_url': font_url,
        'qr_code_base64': qr_code_base64,
    })

    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf = html.write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'filename="invoice_{invoice.number}.pdf"'
    return response



from django.template.loader import render_to_string
from django.templatetags.static import static
from weasyprint import HTML
from django.http import HttpResponse
from .models import SalesReturn
from core.utils import get_company
from .utils.zatca_qr import generate_invoice_qr_base64


def sales_return_invoice_pdf_view(request, pk):
    return_invoice = SalesReturn.objects.get(pk=pk)
    font_url = request.build_absolute_uri(static('fonts/Amiri-Regular.ttf'))
    company = get_company()

    # توليد QR Code Base64
    qr_code_base64 = generate_invoice_qr_base64(
        seller_name=company.name,
        vat_number=company.tax_number,
        timestamp=str(return_invoice.date),
        invoice_total=str(return_invoice.total_amount),
        vat_total=str(return_invoice.tax_amount)
    )

    html_string = render_to_string('sales/sales_return_invoice_pdf.html', {
        'return_invoice': return_invoice,
        'company': company,
        'font_url': font_url,
        'qr_code_base64': qr_code_base64,
    })

    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf = html.write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'filename="sales_return_invoice_{return_invoice.number}.pdf"'
    return response


###########################################################################################################
# reports/views.py
from django.shortcuts import render
from django.db.models import Sum
from sales.models import SalesInvoice, SalesRepresentative
from purchases.models import PurchaseInvoice
from django import forms
from datetime import date
import io
import pandas as pd
from django.http import HttpResponse





class RepFilterForm(forms.Form):
    start_date = forms.DateField(label="من تاريخ", widget=forms.DateInput(attrs={"type": "date"}), initial=date(date.today().year, 1, 1))
    end_date = forms.DateField(label="إلى تاريخ", widget=forms.DateInput(attrs={"type": "date"}), initial=date.today())
    rep = forms.ModelChoiceField(queryset=SalesRepresentative.objects.all(), required=False, label="المندوب")

def rep_sales_purchases_report(request):
    form = RepFilterForm(request.GET or None)
    results = []

    if form.is_valid():
        start = form.cleaned_data["start_date"]
        end = form.cleaned_data["end_date"]
        rep = form.cleaned_data["rep"]

        # جميع المندوبين الذين ظهروا في المبيعات أو المشتريات
        rep_ids_sales = SalesInvoice.objects.filter(date__range=(start, end)).values_list('sales_rep_id', flat=True)
        rep_ids_purchases = PurchaseInvoice.objects.filter(date__range=(start, end)).values_list('sales_rep_id', flat=True)
        all_rep_ids = set(rep_ids_sales) | set(rep_ids_purchases)

        reps = SalesRepresentative.objects.filter(id__in=all_rep_ids)
        if rep:
            reps = reps.filter(id=rep.id)


        for rep in reps:
            # المبيعات
            
            sales_total = SalesInvoice.objects.filter(
                date__range=(start, end),
                sales_rep=rep
            ).aggregate(total=Sum('total_with_tax_value'))['total'] or 0

            # المشتريات
            purchase_total = PurchaseInvoice.objects.filter(
                date__range=(start, end),
                sales_rep=rep
            ).aggregate(total=Sum('total_with_tax_value'))['total'] or 0

            results.append({
                "rep": rep,
                "sales_total": sales_total,
                "purchase_total": purchase_total,
                "net": sales_total - purchase_total
            })

    if request.GET.get("export") == "excel":
        df = pd.DataFrame([{
            "المندوب": row["rep"].name,
            "إجمالي المبيعات": row["sales_total"],
            "إجمالي المشتريات": row["purchase_total"],
            "الفرق (صافي)": row["net"],
        } for row in results])

        excel_file = io.BytesIO()
        df.to_excel(excel_file, index=False)
        excel_file.seek(0)

        response = HttpResponse(
            excel_file.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=rep_sales_purchases_report.xlsx"
 
        return response
    return render(request, "sales/reports/rep_sales_purchases_report.html", {
        "form": form,
        "results": results,
        "title": "تقرير المبيعات والمشتريات حسب المندوب"
    })





# views.py
import datetime
from django.shortcuts import render
from django.db.models import Sum
from sales.models import SalesInvoice, SalesRepresentative, CommissionSlab
from django import forms
from decimal import Decimal
import openpyxl
from django.http import HttpResponse

class CommissionReportForm(forms.Form):
    start_date = forms.DateField(label="من تاريخ", widget=forms.DateInput(attrs={"type": "date"}), initial=datetime.date(datetime.date.today().year, 1, 1))
    end_date = forms.DateField(label="إلى تاريخ", widget=forms.DateInput(attrs={"type": "date"}), initial=datetime.date.today())
    rep = forms.ModelChoiceField(queryset=SalesRepresentative.objects.all(), required=False, label="المندوب")

def calculate_commission(rep, total_sales):
    details = []
    total_commission = Decimal('0')

    if rep.commission_type == 'fixed_percent':
        percent = rep.fixed_commission_percent or 0
        total_commission = (total_sales * percent / 100).quantize(Decimal('0.01'))
        details.append({
            "type": "نسبة ثابتة",
            "percentage": percent,
            "amount": total_commission,
        })

    elif rep.commission_type == 'slabs':
        slabs = rep.commission_slabs.order_by('min_amount')
        remaining_sales = total_sales

        for slab in slabs:
            # تحديد الحد الأعلى لهذه الشريحة
            max_in_slab = min(slab.max_amount, remaining_sales)

            # مقدار المبيعات داخل هذه الشريحة
            if max_in_slab >= slab.min_amount:
                sales_in_slab = max_in_slab - slab.min_amount
                commission = (sales_in_slab * slab.commission_percent / 100).quantize(Decimal('0.01'))
                total_commission += commission

                details.append({
                    "tier": f"من {slab.min_amount} إلى {slab.max_amount} - {slab.commission_percent}%",
                    "sales_in_slab": sales_in_slab,
                    "amount": commission,
                })

            # إذا تجاوزنا أعلى شريحة محددة، نوقف
            if remaining_sales <= slab.max_amount:
                break

    else:
        details.append({"note": "لا توجد سياسة عمولة محددة"})

    return total_commission, details


def commission_report_view(request):
    form = CommissionReportForm(request.GET or None)
    results = []

    if form.is_valid():
        start = form.cleaned_data['start_date']
        end = form.cleaned_data['end_date']
        selected_rep = form.cleaned_data['rep']

        reps = SalesRepresentative.objects.all()
        if selected_rep:
            reps = reps.filter(id=selected_rep.id)

        for rep in reps:
            sales = SalesInvoice.objects.filter(
                date__range=(start, end),
                sales_rep=rep
            ).aggregate(total=Sum('total_before_tax_value'))['total'] or 0

            commission_amount, commission_details = calculate_commission(rep, sales)

            results.append({
                "rep": rep,
                "sales_total": sales,
                "commission": commission_amount,
                "details": commission_details,
            })

        # ✅ تصدير Excel
        if "export" in request.GET:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "عمولات المندوبين"

            ws.append(["المندوب", "إجمالي المبيعات", "العمولة", "تفاصيل العمولة"])
            for row in results:
                details_text = " / ".join(
                    f"{d.get('type', d.get('tier', ''))} = {d['amount']}" for d in row['details']
                )
                ws.append([row['rep'].name, float(row['sales_total']), float(row['commission']), details_text])

            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = 'attachment; filename=commission_report.xlsx'
            wb.save(response)
            return response

    return render(request, "sales/reports/commission_report.html", {
        "form": form,
        "results": results,
        "title": "تقرير عمولات المندوبين",
    })
