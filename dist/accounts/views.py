from django.shortcuts import render
from django.http import HttpResponse
from .models import Account,FinancialStatement



def accounts_reports_home(request):
    return render(request, "accounts/reports/reports_home.html")


#----------------------------------------------------------------------------------

# accounts/views.py

from django.shortcuts import render
from django.db.models import Q, Sum
from .models import Account, JournalEntryLine
from datetime import datetime
from decimal import Decimal

def get_all_descendant_accounts(account):
    """جلب كل الحسابات التابعة لهذا الحساب (مباشرة وغير مباشرة)"""
    descendants = set()
    stack = [account]
    while stack:
        node = stack.pop()
        children = Account.objects.filter(parent=node)
        for child in children:
            descendants.add(child.id)
            stack.append(child)
    return list(descendants)



def ledger_report(request):
    accounts = Account.objects.order_by('code')
    selected_account_id = request.GET.get('account')
    date_from = request.GET.get('from', '2026-01-01')
    date_to = request.GET.get('to', datetime.today().strftime('%Y-%m-%d'))

    lines = []
    selected_account = None
    opening_debit = Decimal('0.00')
    opening_credit = Decimal('0.00')
    account_ids = []
    cumulative_lines = []
    total_debit = Decimal('0.00')
    total_credit = Decimal('0.00')

    if selected_account_id:
        selected_account = Account.objects.get(id=selected_account_id)
        account_ids = [selected_account.id] + get_all_descendant_accounts(selected_account)

        opening = JournalEntryLine.objects.filter(
            account_id__in=account_ids,
            journal_entry__date__lt=date_from
        ).aggregate(
            debit_sum=Sum('debit'),
            credit_sum=Sum('credit')
        )
        opening_debit = opening['debit_sum'] or Decimal('0.00')
        opening_credit = opening['credit_sum'] or Decimal('0.00')

        opening_balance = opening_debit - opening_credit

        # الحركات داخل الفترة
        lines = JournalEntryLine.objects.filter(
            account_id__in=account_ids,
            journal_entry__date__range=[date_from, date_to]
        ).select_related('journal_entry', 'account').order_by('journal_entry__date', 'id')

        cumulative = opening_balance
        for line in lines:
            cumulative += (line.debit or 0) - (line.credit or 0)
            total_debit += line.debit or 0
            total_credit += line.credit or 0
            cumulative_lines.append({
                'date': line.journal_entry.date,
                'entry_id': line.journal_entry.id,
                'account': line.account,
                'description': line.description,
                'debit': line.debit,
                'credit': line.credit,
                'cumulative': cumulative,
            })

    return render(request, 'accounts/reports/ledger_report.html', {
        'accounts': accounts,
        'selected_account': selected_account,
        'lines': cumulative_lines,
        'date_from': date_from,
        'date_to': date_to,
        'opening_debit': opening_debit,
        'opening_credit': opening_credit,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'opening_balance': opening_debit - opening_credit,
        'closing_balance': (opening_debit - opening_credit) + (total_debit - total_credit),
    })



#----------------------------------------------------------------------------------
from django.shortcuts import render
from datetime import datetime, date
from accounts.models import Account, JournalEntryLine
from .models import OpeningBalanceItem
from collections import defaultdict
from django.db.models import Sum

def trial_balance_view(request):
    default_from = date(2025, 1, 1)
    default_to = date(2025, 12, 31)

    date_from = request.GET.get('from')
    date_to = request.GET.get('to')
    level_filter = request.GET.get("level")

    if not date_from:
        date_from = default_from
    else:
        date_from = datetime.strptime(date_from, "%Y-%m-%d").date()

    if not date_to:
        date_to = default_to
    else:
        date_to = datetime.strptime(date_to, "%Y-%m-%d").date()

    accounts = Account.objects.all().order_by("code")
    accounts_by_id = {acc.id: acc for acc in accounts}
    children_map = defaultdict(list)

    for acc in accounts:
        if acc.parent_id:
            children_map[acc.parent_id].append(acc.id)

    # احتساب الرصيد والحركات
    for acc in accounts:
        acc.opening = acc.debit = acc.credit = acc.final = 0

        if acc.level in [4, 5]:
            opening = OpeningBalanceItem.objects.filter(account=acc).aggregate(
                debit=Sum("debit"), credit=Sum("credit")
            )
            acc.opening = (opening["debit"] or 0) - (opening["credit"] or 0)

            movement = JournalEntryLine.objects.filter(
                account=acc,
                journal_entry__date__gte=date_from,
                journal_entry__date__lte=date_to,
            ).exclude(
                journal_entry__description__icontains="افتتاحي"  # ✅ استثناء الأرصدة الافتتاحية
            ).aggregate(
                debit=Sum("debit"),
                credit=Sum("credit")
            )

            acc.debit = movement["debit"] or 0
            acc.credit = movement["credit"] or 0
            acc.final = acc.opening + acc.debit - acc.credit

    # تجميع حسابات الأبناء
    for level in [5, 4, 3, 2, 1]:
        for acc in accounts:
            if acc.level == level:
                for child_id in children_map.get(acc.id, []):
                    child = accounts_by_id[child_id]
                    acc.opening += child.opening
                    acc.debit += child.debit
                    acc.credit += child.credit
                    acc.final += child.final

    # فلترة حسب المستوى المحدد
    if level_filter:
        accounts = [acc for acc in accounts if str(acc.level) == level_filter]

    # إجماليات المستوى المطلوب فقط
    totals = defaultdict(lambda: {"opening": 0, "debit": 0, "credit": 0, "final": 0})
    for acc in accounts:
        lvl = acc.level
        totals[lvl]["opening"] += acc.opening
        totals[lvl]["debit"] += acc.debit
        totals[lvl]["credit"] += acc.credit
        totals[lvl]["final"] += acc.final

    context = {
        "rows": accounts,
        "date_from": date_from.strftime('%Y-%m-%d'),
        "date_to": date_to.strftime('%Y-%m-%d'),
        "selected_level": level_filter,
        "totals": dict(totals),
    }

    # تصدير Excel
    if request.GET.get("export") == "excel":
        import csv
        from django.http import HttpResponse

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="trial_balance.csv"'
        writer = csv.writer(response)
        writer.writerow(["الكود", "اسم الحساب", "رصيد أول المدة", "مدين", "دائن", "رصيد آخر المدة"])

        for acc in accounts:
            writer.writerow([acc.code, acc.name, acc.opening, acc.debit, acc.credit, acc.final])

        return response

    return render(request, 'accounts/reports/trial_balance.html', context)

###########################################################################################################

from django.shortcuts import render, redirect
from .models import Account, OpeningBalance
from django.db.models import Sum

def opening_balances_view(request):
    accounts = Account.objects.filter(level=3).order_by('code')

    # حفظ القيم الجديدة إن وجدت
    if request.method == 'POST':
        for account in accounts:
            debit = request.POST.get(f"debit_{account.id}", '0') or '0'
            credit = request.POST.get(f"credit_{account.id}", '0') or '0'
            OpeningBalance.objects.update_or_create(
                account=account,
                defaults={
                    'debit': debit,
                    'credit': credit
                }
            )
        return redirect('accounts:opening-balances')

    # جلب الأرصدة الحالية
    balances = {b.account_id: b for b in OpeningBalance.objects.all()}
    totals = OpeningBalance.objects.aggregate(
        total_debit=Sum('debit'),
        total_credit=Sum('credit')
    ) or {}

    return render(request, 'accounts/reports/opening_balances.html', {
        'accounts': accounts,
        'balances': balances,
        'total_debit': totals.get('total_debit', 0),
        'total_credit': totals.get('total_credit', 0),
    })


def reports_home(request):
    return render(request, 'accounts/reports/reports_home.html')


####################################################################################################
from collections import defaultdict
from django.shortcuts import render
from django.db.models import Sum
from accounts.models import Account, JournalEntryLine

def income_statement_view(request):
    date_from = request.GET.get("from", "2026-01-01")
    date_to = request.GET.get("to", "2026-12-31")
    level_filter = request.GET.get("level")
    hide_zero = request.GET.get("hide_zero") == "1"  # ✅ قراءة الفلتر الجديد

    # ✅ جلب الحسابات المرتبطة بـ direction.code = IS
    accounts = Account.objects.filter(direction__code="IS")

    if level_filter:
        accounts = accounts.filter(level=level_filter)

    accounts = accounts.select_related("direction").order_by("direction__name", "code")

    grouped_data = defaultdict(list)
    total_revenue = 0
    total_expense = 0

    for acc in accounts:
        movements = JournalEntryLine.objects.filter(
            account=acc,
            journal_entry__date__range=[date_from, date_to]
        ).aggregate(
            debit=Sum("debit") or 0,
            credit=Sum("credit") or 0
        )

        debit = movements["debit"] or 0
        credit = movements["credit"] or 0
        balance = credit - debit

        # ✅ إذا كان خيار الإخفاء مفعّل ولا يوجد رصيد، تجاهل الحساب
        if hide_zero and balance == 0:
            continue

        direction_name = acc.direction.name if acc.direction else "بدون تصنيف"

        grouped_data[direction_name].append({
            "code": acc.code,
            "name": acc.name,
            "level": acc.level,
            "balance": balance
        })

        if balance >= 0:
            total_revenue += balance
        else:
            total_expense += abs(balance)

    net_income = total_revenue - total_expense

    return render(request, "accounts/reports/income_statement.html", {
        "grouped_data": dict(grouped_data),
        "total_revenue": total_revenue,
        "total_expense": total_expense,
        "net_income": net_income,
        "date_from": date_from,
        "date_to": date_to,
        "selected_level": level_filter,
        "hide_zero": hide_zero,  # ✅ لإعادة تفعيل الخيار في النموذج
    })

#######################################################################################################
from datetime import datetime
from decimal import Decimal
from collections import defaultdict
from django.shortcuts import render
from django.db.models import Sum
from accounts.models import Account, JournalEntryLine

def balance_sheet_view(request):
    date_from = request.GET.get("from", "2026-01-01")
    date_to = request.GET.get("to", "2026-12-31")
    level_filter = request.GET.get("level")
    export_excel = request.GET.get("export") == "excel"

    # جميع الحسابات المرتبطة بالميزانية العمومية
    all_accounts = Account.objects.filter(direction__code="BS").select_related("direction", "parent").order_by("code")

    account_map = {acc.code: acc for acc in all_accounts}
    children_map = defaultdict(list)
    for acc in all_accounts:
        if acc.parent:
            children_map[acc.parent.code].append(acc)

    # حساب الرصيد المبدئي
    balances = {}
    for acc in all_accounts:
        movement = JournalEntryLine.objects.filter(
            account=acc,
            journal_entry__date__range=[date_from, date_to]
        ).aggregate(debit=Sum("debit"), credit=Sum("credit"))
        
        debit = movement["debit"] or 0
        credit = movement["credit"] or 0
        balance = Decimal(debit) - Decimal(credit)

        if acc.direction and acc.direction.code in ["L", "E"]:
            balance = -balance

        balances[acc.code] = balance
        acc.balance = balance

    # تجميع المستويات العليا
    def compute_total(code):
        acc = account_map.get(code)
        total = balances.get(code, Decimal(0))
        for child in children_map.get(code, []):
            total += compute_total(child.code)
        balances[code] = total
        return total

    for acc in all_accounts:
        if acc.level == 1:
            compute_total(acc.code)

    # ⬇️ تعريف المتغيرات المهمة قبل الاستخدام
    rows = []
    totals = defaultdict(Decimal)
    level_totals = defaultdict(Decimal)

    # إعداد الجدول الأساسي من الحسابات
    for acc in all_accounts:
        acc.balance = balances.get(acc.code, Decimal(0))
        rows.append({
            "code": acc.code,
            "name": acc.name,
            "level": acc.level,
            "statement_type": acc.direction.name if acc.direction else "غير مصنف",
            "balance": acc.balance,
        })

        if acc.direction:
            totals[acc.direction.code] += acc.balance
        level_totals[acc.level] += acc.balance

    # ✅ حساب صافي الربح / الخسارة من حسابات قائمة الدخل (direction = IS)
    is_movement = JournalEntryLine.objects.filter(
        account__direction__code="IS",
        journal_entry__date__range=[date_from, date_to]
    ).aggregate(
        debit=Sum("debit") or 0,
        credit=Sum("credit") or 0
    )

    revenue = is_movement["credit"] or 0
    expenses = is_movement["debit"] or 0
    net_profit = -(Decimal(revenue) - Decimal(expenses))  # الربح إشارته سالبة

    # ✅ إضافة صف الأرباح/الخسائر لكل مستوى من 1 إلى 4
    for lvl in [1, 2, 3, 4]:
        code = f"9999{lvl}"
        rows.append({
            "code": code,
            "name": "صافي الربح / الخسارة",
            "level": lvl,
            "statement_type": "حقوق الملكية",
            "balance": net_profit,
        })
        totals["E"] += net_profit
        level_totals[lvl] += net_profit

    # ✅ حساب إجمالي الخصوم وحقوق الملكية
    totals["liabilities_and_equity"] = totals["L"] + totals["E"]

    # ✅ فلترة الصفوف حسب المستوى إن وُجد
    if level_filter:
        rows = [row for row in rows if str(row["level"]) == str(level_filter)]

    context = {
        "rows": rows,
        "totals": {
            "assets": totals["A"],
            "liabilities": totals["L"],
            "equity": totals["E"],
            "liabilities_and_equity": totals["liabilities_and_equity"],
        },
        "level_totals": level_totals,
        "date_from": date_from,
        "date_to": date_to,
        "selected_level": level_filter,
        "net_profit": net_profit,
    }

    return render(request, "accounts/reports/balance_sheet.html", context)


####################################################################################################################
from django.shortcuts import render
from django.db.models import Sum
from accounts.models import JournalEntryLine
from datetime import datetime
import pandas as pd
from django.http import HttpResponse

def cash_flow_report_view(request):
    start_date = request.GET.get('start_date', '2026-01-01')
    end_date = request.GET.get('end_date', '2026-12-31')

    activities = {
        'operating': {'title': 'الأنشطة التشغيلية', 'accounts': [], 'total': 0},
        'investing': {'title': 'الأنشطة الاستثمارية', 'accounts': [], 'total': 0},
        'financing': {'title': 'الأنشطة التمويلية', 'accounts': [], 'total': 0},
    }

    lines = JournalEntryLine.objects.filter(
        journal_entry__date__range=(start_date, end_date),
        account__group__activity_type__in=activities.keys()
    ).values(
        'account__id',
        'account__code',
        'account__name',
        'account__group__activity_type'
    ).annotate(
        debit_sum=Sum('debit'),
        credit_sum=Sum('credit')
    ).order_by('account__code')

    for line in lines:
        activity = line['account__group__activity_type']
        net_cash = line['credit_sum'] - line['debit_sum']

        activities[activity]['accounts'].append({
            'code': line['account__code'],
            'name': line['account__name'],
            'debit': line['debit_sum'],
            'credit': line['credit_sum'],
            'net': net_cash
        })

        activities[activity]['total'] += net_cash

    # تصدير إلى Excel
    if request.GET.get("export") == "excel":
        rows = []
        for key, section in activities.items():
            rows.append([section['title'], '', '', '', ''])
            for acc in section['accounts']:
                rows.append([
                    acc['code'], acc['name'], acc['debit'], acc['credit'], acc['net']
                ])
            rows.append(['الإجمالي', '', '', '', section['total']])
            rows.append(['', '', '', '', ''])

        df = pd.DataFrame(rows, columns=['كود الحساب', 'اسم الحساب', 'مدين', 'دائن', 'صافي التدفق'])
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="cash_flow_report.xlsx"'
        df.to_excel(response, index=False)
        return response

    return render(request, 'accounts/reports/cash_flow_report.html', {
        'activities': activities,
        'start_date': start_date,
        'end_date': end_date,
    })


#__________________________________________________________________________________________________

from django.shortcuts import render
from django.db.models import Sum
from accounts.models import Account, JournalEntryLine, FinancialStatementType
from django.http import HttpResponse
import pandas as pd
def financial_statement_report_view(request):
    start_date = request.GET.get("start_date", "2026-01-01")
    end_date = request.GET.get("end_date", "2026-12-31")
    selected_code = request.GET.get("statement_type")
    selected_level = request.GET.get("level")

    # جلب نوع القائمة المالية
    statement_type = None
    if selected_code:
        try:
            statement_type = FinancialStatementType.objects.get(code=selected_code)
        except FinancialStatementType.DoesNotExist:
            statement_type = None

    # جلب الحسابات المرتبطة بالقائمة المالية
    accounts = Account.objects.all()
    if statement_type:
        accounts = accounts.filter(statement_type=statement_type)

    account_map = {acc.id: acc for acc in accounts}
    balances = {}

    # جمع الحركات لكل حساب
    for acc in accounts:
        lines = JournalEntryLine.objects.filter(
            account=acc,
            journal_entry__date__range=(start_date, end_date)
        ).aggregate(
            debit=Sum('debit'),
            credit=Sum('credit')
        )
        debit = lines['debit'] or 0
        credit = lines['credit'] or 0
        direction = acc.direction.code if acc.direction else 'debit'
        balance = debit - credit if direction == 'debit' else credit - debit

        # تحديث الحساب ورصيده
        if acc.id not in balances:
            balances[acc.id] = {
                'account': acc,
                'debit': 0,
                'credit': 0,
                'balance': 0
            }

        balances[acc.id]['debit'] += debit
        balances[acc.id]['credit'] += credit
        balances[acc.id]['balance'] += balance

        # ✅ صعودًا إلى الآباء وجمع القيم
        parent = acc.parent
        while parent:
            if parent.id not in balances:
                balances[parent.id] = {
                    'account': parent,
                    'debit': 0,
                    'credit': 0,
                    'balance': 0
                }
            balances[parent.id]['debit'] += debit
            balances[parent.id]['credit'] += credit
            balances[parent.id]['balance'] += balance
            parent = parent.parent

    final_data = []
    total_debit = total_credit = total_balance = 0

    for b in balances.values():
        acc = b['account']
        final_data.append({
            'code': acc.code,
            'name': acc.name,
            'level': acc.level,
            'debit': b['debit'],
            'credit': b['credit'],
            'balance': b['balance'],
        })

        # ✅ حساب الإجمالي فقط من الحسابات الطرفية
        if not acc.children.exists():
            total_debit += b['debit']
            total_credit += b['credit']
            total_balance += b['balance']

    # التصدير إلى Excel
    if request.GET.get("export") == "excel":
        df = pd.DataFrame(final_data)
        df.loc['الإجمالي'] = ["", "الإجمالي", "", total_debit, total_credit, total_balance]
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = f"{selected_code or 'statement'}_report.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        df.to_excel(response, index=False)
        return response

    statement_types = FinancialStatementType.objects.all()

    return render(request, "accounts/reports/financial_statement_report.html", {
        "accounts": sorted(final_data, key=lambda x: x['code']),
        "statement_types": statement_types,
        "selected_code": selected_code,
        "start_date": start_date,
        "end_date": end_date,
        "level": selected_level,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "total_balance": total_balance,
    })




############################################################################################################
from django.shortcuts import render
from django.db.models import Sum
from sales.models import SalesInvoice, SalesReturn
from purchases.models import PurchaseInvoice, PurchaseReturn
import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse

# دالة مساعدة لحساب المجاميع الثلاثة (قبل الضريبة، الضريبة، بعد الضريبة)
def sum_fields(queryset):
    return {
        'before_tax': queryset.aggregate(s=Sum('total_before_tax_value'))['s'] or 0,
        'tax': queryset.aggregate(s=Sum('total_tax_value'))['s'] or 0,
        'with_tax': queryset.aggregate(s=Sum('total_with_tax_value'))['s'] or 0,
    }


def vat_report_view(request):
    # تحديد فترة التقرير مع قيم افتراضية
    date_from = request.GET.get('date_from', '2026-01-01')
    date_to = request.GET.get('date_to', '2026-12-31')

    # استعلام الفواتير والمردودات فقط للفواتير المرحلة (is_posted=True)
    sales_qs = SalesInvoice.objects.filter(date__range=[date_from, date_to], is_posted=True)
    sales_return_qs = SalesReturn.objects.filter(date__range=[date_from, date_to], is_posted=True)
    purchase_qs = PurchaseInvoice.objects.filter(date__range=[date_from, date_to], is_posted=True)
    purchase_return_qs = PurchaseReturn.objects.filter(date__range=[date_from, date_to], is_posted=True)

    # حساب الإجماليات
    sales_totals = sum_fields(sales_qs)
    sales_return_totals = sum_fields(sales_return_qs)
    purchase_totals = sum_fields(purchase_qs)
    purchase_return_totals = sum_fields(purchase_return_qs)

    # حساب الفرق في الضريبة المستحقة
    total_sales_vat = sales_totals['tax'] - sales_return_totals['tax']
    total_purchase_vat = purchase_totals['tax'] - purchase_return_totals['tax']
    vat_difference = total_sales_vat - total_purchase_vat

    # تمرير البيانات إلى القالب
    context = {
        # البيانات التفصيلية
        'sales': sales_qs,
        'sales_returns': sales_return_qs,
        'purchases': purchase_qs,
        'purchase_returns': purchase_return_qs,

        # الإجماليات
        'sales_totals': sales_totals,
        'sales_return_totals': sales_return_totals,
        'purchase_totals': purchase_totals,
        'purchase_return_totals': purchase_return_totals,

        # الفرق النهائي
        'total_sales_vat': total_sales_vat,
        'total_purchase_vat': total_purchase_vat,
        'vat_difference': vat_difference,

        # الفلاتر المستخدمة
        'date_from': date_from,
        'date_to': date_to,
    }

    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "VAT Report"

        # العنوان
        ws.merge_cells("A1:F1")
        ws["A1"] = "تقرير ضريبة القيمة المضافة"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:F2")
        ws["A2"] = f"الفترة من {date_from} إلى {date_to}"
        ws["A2"].alignment = Alignment(horizontal="center")

        row = 4

        # فواتير المبيعات
        ws.append(["🔹 فواتير المبيعات", "", "", "", "", ""])
        ws.append(["رقم الفاتورة", "تاريخ", "قبل الضريبة", "الضريبة", "الإجمالي", "ملاحظة"])
        for s in sales_qs:
            ws.append([s.number, s.date.strftime('%Y-%m-%d'), float(s.total_before_tax_value),
                       float(s.total_tax_value), float(s.total_with_tax_value), "فاتورة مبيعات"])

        # مردودات المبيعات
        ws.append([])
        ws.append(["🔸 مردودات المبيعات", "", "", "", "", ""])
        ws.append(["رقم المردود", "تاريخ", "قبل الضريبة", "الضريبة", "الإجمالي", "ملاحظة"])
        for sr in sales_return_qs:
            ws.append([sr.number, sr.date.strftime('%Y-%m-%d'), float(sr.total_before_tax_value),
                       float(sr.total_tax_value), float(sr.total_with_tax_value), "مردود مبيعات"])

        # فواتير المشتريات
        ws.append([])
        ws.append(["🔹 فواتير المشتريات", "", "", "", "", ""])
        ws.append(["رقم الفاتورة", "تاريخ", "قبل الضريبة", "الضريبة", "الإجمالي", "ملاحظة"])
        for p in purchase_qs:
            ws.append([p.number, p.date.strftime('%Y-%m-%d'), float(p.total_before_tax_value),
                       float(p.total_tax_value), float(p.total_with_tax_value), "فاتورة مشتريات"])

        # مردودات المشتريات
        ws.append([])
        ws.append(["🔸 مردودات المشتريات", "", "", "", "", ""])
        ws.append(["رقم المردود", "تاريخ", "قبل الضريبة", "الضريبة", "الإجمالي", "ملاحظة"])
        for pr in purchase_return_qs:
            ws.append([pr.number, pr.date.strftime('%Y-%m-%d'), float(pr.total_before_tax_value),
                       float(pr.total_tax_value), float(pr.total_with_tax_value), "مردود مشتريات"])

        # ملخص
        ws.append([])
        ws.append(["📌 ملخص", "", "", "", "", ""])
        ws.append(["إجمالي ضريبة المبيعات", "", "", float(sales_totals['tax'])])
        ws.append(["إجمالي مردودات المبيعات", "", "", float(sales_return_totals['tax'])])
        ws.append(["ضريبة المبيعات الصافية", "", "", float(total_sales_vat)])
        ws.append(["إجمالي ضريبة المشتريات", "", "", float(purchase_totals['tax'])])
        ws.append(["إجمالي مردودات المشتريات", "", "", float(purchase_return_totals['tax'])])
        ws.append(["ضريبة المشتريات الصافية", "", "", float(total_purchase_vat)])
        ws.append(["📌 الضريبة المستحقة", "", "", float(vat_difference)])

        # الاستجابة
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=vat_report.xlsx'
        wb.save(response)
        return response

    return render(request, 'accounts/reports/vat_report.html', context)
