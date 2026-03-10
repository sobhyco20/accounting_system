from django.shortcuts import render
from django.utils.dateparse import parse_date
from .models import PurchaseInvoice, PurchaseReturn
from django.db.models import Sum
from datetime import datetime
from django.http import JsonResponse
from .models import PurchaseInvoice

def load_supplier_invoices(request):
    supplier_id = request.GET.get('supplier')
    invoices = PurchaseInvoice.objects.filter(supplier_id=supplier_id, is_posted=True)

    data = []
    for invoice in invoices:
        data.append({
            'id': invoice.id,
            'invoice_number': str(invoice),  # أو: invoice.number
            'total': float(invoice.total_with_tax_value) if invoice.total_with_tax_value else 0,
        })

    return JsonResponse(data, safe=False)


######################################################################################################################
from django.shortcuts import render

def purchase_reports(request):
    return render(request, 'reports/reports_home.html')


from django.shortcuts import render
from .models import PurchaseInvoice, PurchaseReturn

def purchase_report(request):
    date_from = request.GET.get('date_from', '2026-01-01')
    date_to = request.GET.get('date_to', '2026-12-31')

    purchases = PurchaseInvoice.objects.filter(date__range=[date_from, date_to]).prefetch_related('items')
    returns = PurchaseReturn.objects.filter(date__range=[date_from, date_to]).prefetch_related('details')

    # حساب المجاميع لكل فاتورة على حدة
    for invoice in purchases:
        invoice.total_before_tax = sum(item.total_before_tax for item in invoice.items.all())

    for ret in returns:
        ret.total_before_tax = sum(item.total_before_tax for item in ret.details.all())

    # إجماليات عامة
    total_purchase_before_tax = sum(inv.total_before_tax for inv in purchases)
    total_purchase_tax = sum(inv.total_tax_value for inv in purchases)
    total_purchase_after_tax = sum(inv.total_with_tax_value for inv in purchases)

    total_return_before_tax = sum(ret.total_before_tax for ret in returns)
    total_return_tax = sum(ret.total_tax_value for ret in returns)
    total_return_after_tax = sum(ret.total_with_tax_value for ret in returns)

    net_total_before_tax = total_purchase_before_tax - total_return_before_tax
    net_total_tax = total_purchase_tax - total_return_tax
    net_total_after_tax = total_purchase_after_tax - total_return_after_tax

    context = {
        'purchases': purchases,
        'returns': returns,
        'totals': {
            'total_before_tax': total_purchase_before_tax,
            'tax_amount': total_purchase_tax,
            'total_with_tax': total_purchase_after_tax,
        },
        'returns_totals': {
            'total_before_tax': total_return_before_tax,
            'tax_amount': total_return_tax,
            'total_with_tax': total_return_after_tax,
        },
        'net_totals': {
            'total_before_tax': net_total_before_tax,
            'tax_amount': net_total_tax,
            'total_with_tax': net_total_after_tax,
        },
        'total_purchase_before_tax': total_purchase_before_tax,
        'total_purchase_tax': total_purchase_tax,
        'total_purchase_after_tax': total_purchase_after_tax,
        'total_return_after_tax': total_return_after_tax,
        'net_total': net_total_after_tax,
        'request': request,
    }

    return render(request, 'reports/purchase_report.html', context)

################################################################################
from django.shortcuts import render
from django.http import HttpResponse
from .models import Supplier, PurchaseInvoice, PurchaseReturn, SupplierPayment
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO

def supplier_ledger_report(request):
    suppliers = Supplier.objects.all()
    supplier_id = request.GET.get("supplier")
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")
    export = request.GET.get("export")

    ledger = []
    opening_balance = 0

    if supplier_id and date_from and date_to:
        supplier = Supplier.objects.get(id=supplier_id)
        opening_balance = supplier.opening_credit - supplier.opening_debit

        purchases_before = PurchaseInvoice.objects.filter(supplier=supplier, date__lt=date_from)
        returns_before = PurchaseReturn.objects.filter(supplier=supplier, date__lt=date_from)
        payments_before = SupplierPayment.objects.filter(supplier=supplier, date__lt=date_from)

        total_purchases_before = sum(
            sum(item.total_before_tax + item.tax_amount for item in inv.items.all())
            for inv in purchases_before
        )
        total_returns_before = sum(
            sum(item.total_before_tax + item.tax_amount for item in ret.details.all())
            for ret in returns_before
        )
        total_payments_before = sum(pay.amount for pay in payments_before)

        opening_balance += total_returns_before - total_purchases_before + total_payments_before

        movements = []

        purchases = PurchaseInvoice.objects.filter(supplier=supplier, date__gte=date_from, date__lte=date_to)
        for inv in purchases:
            total_invoice = sum(item.total_before_tax + item.tax_amount for item in inv.items.all())
            movements.append({
                'date': inv.date,
                'description': f"فاتورة مشتريات رقم {inv.number}",
                'debit': 0,
                'credit': total_invoice
            })

        returns = PurchaseReturn.objects.filter(supplier=supplier, date__gte=date_from, date__lte=date_to)
        for ret in returns:
            total_return = sum(item.total_before_tax + item.tax_amount for item in ret.details.all())
            movements.append({
                'date': ret.date,
                'description': f"مرتجع مشتريات رقم {ret.number}",
                'debit': total_return,
                'credit': 0
            })

        payments = SupplierPayment.objects.filter(supplier=supplier, date__gte=date_from, date__lte=date_to)
        for pay in payments:
            movements.append({
                'date': pay.date,
                'description': f"سداد رقم {pay.number}",
                'debit': pay.amount,
                'credit': 0
            })

        movements = sorted(movements, key=lambda x: x['date'])

        balance = opening_balance

        ledger.append({
            'date': "رصيد أول المدة",
            'description': '',
            'debit': opening_balance if opening_balance > 0 else 0,
            'credit': abs(opening_balance) if opening_balance < 0 else 0,
            'balance': balance
        })

        for move in movements:
            balance += move['debit'] - move['credit']
            move['balance'] = balance
            ledger.append(move)

        if export == 'excel':
            return export_to_excel(ledger, supplier, opening_balance)

    context = {
        'title': "كشف حساب مورد",
        'suppliers': suppliers,
        'ledger': ledger,
        'opening_balance': opening_balance,
        'selected_supplier': int(supplier_id) if supplier_id else None,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, "reports/supplier_ledger_report.html", context)

def export_to_excel(ledger, supplier, opening_balance):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "كشف حساب"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal="center")
    bold_font = Font(bold=True)

    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = f"كشف حساب المورد: {supplier.name}"
    cell.font = Font(bold=True, size=14)
    cell.alignment = center_alignment

    ws.merge_cells('A2:E2')
    cell = ws['A2']
    cell.value = f"رصيد أول المدة: {opening_balance:,.2f}"
    cell.font = Font(bold=True, size=12)
    cell.alignment = center_alignment

    headers = ["التاريخ", "البيان", "مدين", "دائن", "الرصيد التراكمي"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    for row_num, row in enumerate(ledger, 5):
        ws.cell(row=row_num, column=1, value=str(row['date'])).alignment = center_alignment
        ws.cell(row=row_num, column=2, value=row['description']).alignment = Alignment(horizontal="right")
        ws.cell(row=row_num, column=3, value=row['debit']).number_format = '#,##0.00'
        ws.cell(row=row_num, column=4, value=row['credit']).number_format = '#,##0.00'
        ws.cell(row=row_num, column=5, value=row['balance']).number_format = '#,##0.00'

        for col in range(1, 6):
            ws.cell(row=row_num, column=col).border = thin_border

    current_balance = ledger[-1]['balance'] if ledger else opening_balance

    summary_row = len(ledger) + 6
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=4)
    ws.cell(row=summary_row, column=1, value="الرصيد الحالي").font = bold_font
    ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=summary_row, column=5, value=current_balance).font = bold_font
    ws.cell(row=summary_row, column=5).number_format = '#,##0.00'
    ws.cell(row=summary_row, column=5).alignment = center_alignment

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value) if cell.value is not None else "") for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length + 5

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename=ledger_{supplier.name}.xlsx'

    buffer = BytesIO()
    wb.save(buffer)
    response.write(buffer.getvalue())
    return response
########################################################################################################################
from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum, F
from purchases.models import PurchaseInvoiceItem
from inventory.models import Product
from purchases.models import Supplier
import openpyxl

def purchases_by_supplier_product(request):
    # التواريخ الافتراضية
    default_start_date = "2026-01-01"
    default_end_date = "2026-12-31"

    # جلب الفلاتر من GET
    start_date = request.GET.get('start_date') or default_start_date
    end_date = request.GET.get('end_date') or default_end_date
    supplier_id = request.GET.get('supplier')
    product_id = request.GET.get('product')

    # الاستعلام الأساسي
    items = PurchaseInvoiceItem.objects.select_related(
        'invoice__supplier', 'product'
    ).filter(
        invoice__date__gte=start_date,
        invoice__date__lte=end_date
    )

    if supplier_id:
        items = items.filter(invoice__supplier__id=supplier_id)
    if product_id:
        items = items.filter(product__id=product_id)

    grouped_data = items.values(
        supplier_name=F('invoice__supplier__name'),
        product_name=F('product__name')
    ).annotate(
        total_quantity=Sum('quantity'),
        total_before_tax=Sum('total_before_tax'),
        total_tax=Sum('tax_amount'),
        total_with_tax=Sum('total_with_tax')
    ).order_by('supplier_name', 'product_name')

    # تصدير إلى Excel
    if 'export' in request.GET:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "تقرير المشتريات"
        headers = ["المورد", "الصنف", "الكمية", "الإجمالي قبل الضريبة", "الضريبة", "الإجمالي مع الضريبة"]
        ws.append(headers)

        for row in grouped_data:
            ws.append([
                row['supplier_name'],
                row['product_name'],
                float(row['total_quantity']),
                float(row['total_before_tax']),
                float(row['total_tax']),
                float(row['total_with_tax']),
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"purchases_filtered_{start_date}_to_{end_date}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

    # جلب كل الموردين والمنتجات للقائمة المنسدلة
    suppliers = Supplier.objects.all()
    products = Product.objects.all()

    context = {
        'grouped_data': grouped_data,
        'start_date': start_date,
        'end_date': end_date,
        'supplier_id': supplier_id,
        'product_id': product_id,
        'suppliers': suppliers,
        'products': products,
    }
    return render(request, 'reports/purchases_by_supplier_product.html', context)


###########################################################################################################
from django.shortcuts import render
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from purchases.models import PurchaseInvoice, Supplier
import openpyxl

def purchases_by_supplier_report(request):
    suppliers = Supplier.objects.all()
    supplier_id = request.GET.get("supplier")
    date_from = request.GET.get("date_from") or "2026-01-01"
    date_to = request.GET.get("date_to") or "2026-12-31"

    purchases_qs = PurchaseInvoice.objects.select_related('supplier').prefetch_related('items') \
        .filter(date__gte=parse_date(date_from), date__lte=parse_date(date_to))

    if supplier_id:
        purchases_qs = purchases_qs.filter(supplier_id=supplier_id)

    # الحسابات لكل فاتورة
    for inv in purchases_qs:
        inv.total_before_tax_value = sum(i.total_before_tax for i in inv.items.all())
        inv.total_tax_value = sum(i.tax_amount for i in inv.items.all())
        inv.total_with_tax_value = inv.total_before_tax_value + inv.total_tax_value

    # الإجماليات
    totals = {
        'total_before_tax': sum(inv.total_before_tax_value for inv in purchases_qs),
        'tax_amount': sum(inv.total_tax_value for inv in purchases_qs),
        'total_with_tax': sum(inv.total_with_tax_value for inv in purchases_qs),
    }

    # تجميع البيانات حسب المورد
    chart_data = {}
    for inv in purchases_qs:
        name = inv.supplier.name
        chart_data[name] = chart_data.get(name, 0) + inv.total_with_tax_value

    # نحولها إلى قائمتين لإرسالها للقالب
    chart_labels = list(chart_data.keys())
    chart_values = list(chart_data.values())

    # Excel export
    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "تقرير المشتريات"

        ws.append(["رقم الفاتورة", "تاريخ", "المورد", "قبل الضريبة", "الضريبة", "الإجمالي"])
        for inv in purchases_qs:
            ws.append([
                inv.number,
                inv.date.strftime('%Y-%m-%d'),
                inv.supplier.name,
                round(inv.total_before_tax_value, 2),
                round(inv.total_tax_value, 2),
                round(inv.total_with_tax_value, 2),
            ])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename=purchases_by_supplier.xlsx'
        wb.save(response)
        return response

    context = {
        'purchases': purchases_qs,
        'totals': totals,
        'suppliers': suppliers,
        'selected_supplier': int(supplier_id) if supplier_id else None,
        'date_from': date_from,
        'date_to': date_to,
        'chart_labels': chart_labels,
        'chart_values': chart_values,
    }

    return render(request, 'reports/purchases_by_supplier_report.html', context)




# purchases/views.py
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.templatetags.static import static
from weasyprint import HTML
from django.http import HttpResponse
from core.utils import get_company
from .models import PurchaseReturn  # ✅

def purchase_return_invoice_pdf_view(request, pk):
    return_invoice = get_object_or_404(PurchaseReturn, pk=pk)
    company = get_company()
    font_url = request.build_absolute_uri(static('fonts/Amiri-Regular.ttf'))

    html_string = render_to_string('purchases/purchase_return_invoice_pdf.html', {
        'return_invoice': return_invoice,
        'company': company,
        'font_url': font_url,
    })

    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf = html.write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'filename="purchase_return_{return_invoice.number}.pdf"'
    return response

from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.templatetags.static import static
from django.template.loader import render_to_string
from weasyprint import HTML
from .models import PurchaseInvoice
from core.utils import get_company


def purchase_invoice_pdf_view(request, invoice_id):
    invoice = get_object_or_404(PurchaseInvoice, id=invoice_id)
    company = get_company()
    font_url = request.build_absolute_uri(static('fonts/Amiri-Regular.ttf'))

    html_string = render_to_string('purchases/purchase_invoice_pdf.html', {
        'invoice': invoice,
        'company': company,
        'font_url': font_url,
    })

    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf = html.write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'filename="purchase_invoice_{invoice.number}.pdf"'
    return response
