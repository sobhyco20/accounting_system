# portal/api.py
import json
from decimal import Decimal

from django.http import JsonResponse
from django.views.decorators.http import require_GET, require_POST, require_http_methods
from django.contrib.admin.views.decorators import staff_member_required
from django.db import transaction
from django.db.models import Q, Sum
from django.shortcuts import get_object_or_404
from django.utils import timezone

from expenses.models import (
    Period,
    ExpenseCategory, ExpenseItem, ExpenseLine,
    ExpenseBatch
)

from costing.models import Unit, Product, RawMaterial, BillOfMaterial, BOMItem
from inventory.models import StockCount, StockCountLine
from sales.models import SalesSummary, SalesSummaryLine

# اختياري: مشتريات لو موجودة
try:
    from purchases.models import PurchaseSummary, PurchaseSummaryLine
except Exception:
    PurchaseSummary = None
    PurchaseSummaryLine = None


# =========================
# Helpers
# =========================
D0 = Decimal("0")

def _bad(msg: str, status: int = 400):
    return JsonResponse({"ok": False, "error": msg}, status=status)

def _d(v, default="0") -> Decimal:
    try:
        return Decimal(str(v if v is not None else default))
    except Exception:
        return Decimal(str(default))

def _s(v):
    return "" if v is None else str(v)

def _dec(v) -> Decimal:
    try:
        return Decimal(str(v or "0"))
    except Exception:
        return Decimal("0")

def is_period_locked(period) -> bool:
    return bool(getattr(period, "is_closed", False))

def _pick_first_decimal(obj, field_names, default=None):
    """
    يرجع أول قيمة Decimal صالحة من قائمة حقول داخل object.
    """
    for fn in field_names:
        if not hasattr(obj, fn):
            continue
        v = getattr(obj, fn, None)
        if v is None:
            continue
        try:
            return Decimal(str(v))
        except Exception:
            continue
    return default


# =========================
# Costs (RAW / SEMI)
# =========================
def raw_unit_cost(rm: RawMaterial) -> Decimal:
    """
    تكلفة الخام:
    1) cost_per_ingredient_unit
    2) purchase_price_per_storage_unit / storage_to_ingredient_factor
    3) fallback unit_cost/avg_cost/...
    """
    v = _pick_first_decimal(rm, ["cost_per_ingredient_unit"], default=None)
    if v is not None:
        return v

    big_price = _pick_first_decimal(rm, ["purchase_price_per_storage_unit"], default=None)
    factor = _pick_first_decimal(rm, ["storage_to_ingredient_factor"], default=None)
    try:
        if big_price is not None and factor is not None and factor > 0:
            return (big_price / factor)
    except Exception:
        pass

    v2 = _pick_first_decimal(rm, ["unit_cost", "avg_cost", "last_cost", "standard_cost"], default=None)
    if v2 is not None:
        return v2

    return Decimal("0")


# =========================
# Expenses (Categories/Items) APIs
# =========================
@staff_member_required
@require_GET
def exp_categories_list(request):
    rows = list(
        ExpenseCategory.objects.order_by("code").values(
            "id", "code", "name", "nature", "directness", "frequency", "behavior", "is_active"
        )
    )
    return JsonResponse({"ok": True, "rows": rows})


@staff_member_required
@require_http_methods(["POST"])
@transaction.atomic
def exp_categories_action(request):
    action = request.POST.get("action")

    if action == "save":
        cid = request.POST.get("id") or None
        obj = ExpenseCategory.objects.filter(id=cid).first() if cid else ExpenseCategory()

        obj.code = (request.POST.get("code") or "").strip()
        obj.name = (request.POST.get("name") or "").strip()
        obj.nature = request.POST.get("nature") or "OP"
        obj.directness = request.POST.get("directness") or "INDIRECT"
        obj.frequency = request.POST.get("frequency") or "MONTHLY"
        obj.behavior = request.POST.get("behavior") or "FIXED"
        obj.is_active = (request.POST.get("is_active") == "true")

        obj.full_clean()
        obj.save()
        return JsonResponse({"ok": True, "id": obj.id})

    if action == "delete":
        cid = request.POST.get("id")
        if ExpenseItem.objects.filter(category_id=cid).exists():
            return JsonResponse({"ok": False, "error": "لا يمكن حذف التصنيف لوجود مصروفات مرتبطة به."}, status=409)
        ExpenseCategory.objects.filter(id=cid).delete()
        return JsonResponse({"ok": True})

    return JsonResponse({"ok": False, "error": "عملية غير معروفة"}, status=400)


@staff_member_required
@require_GET
def exp_items_list(request):
    rows = []
    qs = ExpenseItem.objects.select_related("category").order_by("code")
    for it in qs:
        rows.append({
            "id": it.id,
            "code": it.code,
            "name": it.name,
            "category_id": it.category_id,
            "category_name": str(it.category) if it.category_id else "",
            "default_amount": str(it.default_amount or "0.00"),
            "is_active": it.is_active,
        })

    cats = list(ExpenseCategory.objects.order_by("code").values("id", "code", "name"))
    return JsonResponse({"ok": True, "rows": rows, "categories": cats})


@staff_member_required
@require_http_methods(["POST"])
@transaction.atomic
def exp_items_action(request):
    action = request.POST.get("action")

    if action == "save":
        iid = request.POST.get("id") or None
        obj = ExpenseItem.objects.filter(id=iid).first() if iid else ExpenseItem()

        obj.code = (request.POST.get("code") or "").strip()
        obj.name = (request.POST.get("name") or "").strip()
        obj.category_id = request.POST.get("category_id") or None
        obj.default_amount = Decimal(request.POST.get("default_amount") or "0")
        obj.is_active = (request.POST.get("is_active") == "true")

        obj.full_clean()
        obj.save()
        return JsonResponse({"ok": True, "id": obj.id})

    if action == "delete":
        iid = request.POST.get("id")
        if ExpenseLine.objects.filter(item_id=iid).exists():
            return JsonResponse({"ok": False, "error": "لا يمكن حذف المصروف لأنه مستخدم داخل مصروفات الفترات."}, status=409)
        ExpenseItem.objects.filter(id=iid).delete()
        return JsonResponse({"ok": True})

    return JsonResponse({"ok": False, "error": "عملية غير معروفة"}, status=400)


def semi_unit_cost(p: Product) -> Decimal:
    """
    تكلفة النصف مصنع:
    1) من المنتج مباشرة
    2) من BOM: unit_cost أو total_cost/produced_qty
    """
    direct = _pick_first_decimal(p, ["unit_cost", "standard_cost", "avg_cost"], default=None)
    if direct is not None:
        return direct

    bom = BillOfMaterial.objects.filter(product=p).order_by("-id").first()
    if bom:
        uc = _pick_first_decimal(bom, ["unit_cost"], default=None)
        if uc is not None:
            return uc

        total = _pick_first_decimal(bom, ["total_cost"], default=None)
        produced = _pick_first_decimal(bom, ["batch_output_quantity", "produced_qty"], default=None)
        try:
            total = Decimal(str(total or "0"))
            produced = Decimal(str(produced or "0"))
            if produced > 0:
                return total / produced
        except Exception:
            pass

    return Decimal("0")


# =========================
# Units API
# =========================
@staff_member_required
@require_GET
def units_list(request):
    rows = list(Unit.objects.order_by("name").values("id", "name", "abbreviation"))
    return JsonResponse({"ok": True, "rows": rows})

@staff_member_required
@require_POST
def units_api(request):
    action = request.POST.get("action")

    if action == "save":
        uid = request.POST.get("id")
        name = (request.POST.get("name") or "").strip()
        abbr = (request.POST.get("abbreviation") or "").strip()

        if not name:
            return _bad("name is required")

        u = Unit.objects.filter(id=uid).first() if uid else Unit()
        if uid and not u:
            return _bad("unit not found", 404)

        u.name = name
        u.abbreviation = abbr
        u.save()
        return JsonResponse({"ok": True, "id": u.id})

    if action == "delete":
        Unit.objects.filter(id=request.POST.get("id")).delete()
        return JsonResponse({"ok": True})

    return _bad("invalid action")


# =========================
# Products API
# =========================
@staff_member_required
@require_GET
def products_list(request):
    """
    type:
      - sell: منتجات بيع (is_sellable=True, is_semi_finished=False)
      - semi: نصف مصنع (is_semi_finished=True)
      - all: الكل
    """
    t = (request.GET.get("type") or "all").strip()
    q = (request.GET.get("q") or "").strip()

    qs = Product.objects.all().select_related("base_unit").order_by("code")

    if t == "sell":
        qs = qs.filter(is_sellable=True, is_semi_finished=False)
    elif t == "semi":
        qs = qs.filter(is_semi_finished=True)

    if q:
        qs = qs.filter(Q(code__icontains=q) | Q(name__icontains=q) | Q(name_en__icontains=q))

    rows = []
    for p in qs[:2000]:
        rows.append({
            "id": p.id,
            "code": p.code,
            "name": p.name,
            "name_en": p.name_en or "",
            "base_unit_id": p.base_unit_id,
            "base_unit_name": getattr(p.base_unit, "name", "") if p.base_unit_id else "",
            "is_sellable": bool(p.is_sellable),
            "is_semi_finished": bool(p.is_semi_finished),
            "selling_price_per_unit": str(getattr(p, "selling_price_per_unit", None) or "0"),
        })

    units = list(Unit.objects.order_by("name").values("id", "name", "abbreviation"))
    return JsonResponse({"ok": True, "rows": rows, "units": units})

@staff_member_required
@require_POST
def products_api(request):
    action = request.POST.get("action")

    if action == "save":
        pid = request.POST.get("id")
        p = Product.objects.filter(id=pid).first() if pid else Product()
        if pid and not p:
            return _bad("product not found", 404)

        p.code = (request.POST.get("code") or "").strip()
        p.name = (request.POST.get("name") or "").strip()
        p.name_en = (request.POST.get("name_en") or "").strip() or None
        p.base_unit_id = request.POST.get("base_unit_id") or None
        p.is_sellable = (request.POST.get("is_sellable") == "1")
        p.is_semi_finished = (request.POST.get("is_semi_finished") == "1")
        p.selling_price_per_unit = _d(request.POST.get("selling_price_per_unit") or "0")

        if not p.code or not p.name or not p.base_unit_id:
            return _bad("code, name, base_unit are required")

        p.save()
        return JsonResponse({"ok": True, "id": p.id})

    if action == "delete":
        Product.objects.filter(id=request.POST.get("id")).delete()
        return JsonResponse({"ok": True})

    return _bad("invalid action")


# =========================
# Raw Materials API
# =========================
def _has_activity(rm: RawMaterial) -> bool:
    rel_names = ["bom_items", "consumption_sources", "salesconsumption", "stockcountline"]
    for rel in rel_names:
        try:
            mgr = getattr(rm, rel)
            if mgr.exists():
                return True
        except Exception:
            pass

    if PurchaseSummaryLine:
        try:
            if PurchaseSummaryLine.objects.filter(raw_material=rm).exists():
                return True
        except Exception:
            pass

    return False

@staff_member_required
@require_GET
def raw_materials_list(request):
    q = (request.GET.get("q") or "").strip()
    qs = RawMaterial.objects.select_related("storage_unit", "ingredient_unit").order_by("sku", "name")
    if q:
        qs = qs.filter(Q(sku__icontains=q) | Q(name__icontains=q))

    rows = []
    for rm in qs[:5000]:
        rows.append({
            "id": rm.id,
            "sku": rm.sku or "",
            "name": rm.name or "",
            "storage_unit_id": rm.storage_unit_id,
            "storage_unit_name": (rm.storage_unit.name if rm.storage_unit_id else ""),
            "ingredient_unit_id": rm.ingredient_unit_id,
            "ingredient_unit_name": (rm.ingredient_unit.name if rm.ingredient_unit_id else ""),
            "factor": str(getattr(rm, "storage_to_ingredient_factor", None) or "0"),
            "big_price": str(getattr(rm, "purchase_price_per_storage_unit", None) or "0"),
            "small_cost": str(raw_unit_cost(rm)),
            "locked": _has_activity(rm),
        })

    units = list(Unit.objects.order_by("name").values("id", "name", "abbreviation"))
    return JsonResponse({"ok": True, "rows": rows, "units": units})

@staff_member_required
@require_POST
def raw_materials_api(request):
    action = request.POST.get("action")

    if action == "save":
        rid = request.POST.get("id")
        rm = RawMaterial.objects.filter(id=rid).first() if rid else RawMaterial()
        if rid and not rm:
            return _bad("raw material not found", 404)

        if rm.id and _has_activity(rm):
            return JsonResponse({"ok": False, "error": "ممنوع الحفظ: هذه المادة عليها حركة."}, status=409)

        rm.sku = (request.POST.get("sku") or "").strip() or None
        rm.name = (request.POST.get("name") or "").strip()

        rm.storage_unit_id = request.POST.get("storage_unit_id") or None
        rm.ingredient_unit_id = request.POST.get("ingredient_unit_id") or None

        rm.storage_to_ingredient_factor = _d(request.POST.get("factor") or "0")
        rm.purchase_price_per_storage_unit = _d(request.POST.get("big_price") or "0")

        if not rm.name:
            return _bad("name is required")

        rm.save()
        return JsonResponse({
            "ok": True,
            "id": rm.id,
            "small_cost": str(raw_unit_cost(rm)),
            "locked": _has_activity(rm),
        })

    if action == "delete":
        rid = request.POST.get("id")
        rm = RawMaterial.objects.filter(id=rid).first()
        if not rm:
            return JsonResponse({"ok": True})

        if _has_activity(rm):
            return JsonResponse({"ok": False, "error": "ممنوع الحذف: هذه المادة عليها حركة."}, status=409)

        rm.delete()
        return JsonResponse({"ok": True})

    return _bad("invalid action")


# =========================
# Periods API (stock toggles)
# =========================
def _has_period_activity(period: Period) -> bool:
    if ExpenseBatch.objects.filter(period=period).exists():
        return True
    if SalesSummary.objects.filter(period=period).exists():
        return True
    if PurchaseSummary and PurchaseSummary.objects.filter(period=period).exists():
        return True
    return False

def _has_next_period_activity(period: Period) -> bool:
    nxt = Period.objects.filter(
        (Q(year=period.year) & Q(month__gt=period.month)) | Q(year__gt=period.year)
    ).order_by("year", "month").first()
    if not nxt:
        return False
    return _has_period_activity(nxt)

@staff_member_required
@require_GET
def api_periods_list(request):
    periods = Period.objects.order_by("-year", "-month")
    rows = []
    for p in periods:
        has_activity = _has_period_activity(p)
        has_next_activity = _has_next_period_activity(p)

        rows.append({
            "id": p.id,
            "year": p.year,
            "month": p.month,
            "label": f"{p.year}-{p.month:02d}",
            "is_closed": bool(p.is_closed),
            "has_activity": bool(has_activity),

            "allow_opening_stock": bool(getattr(p, "allow_opening_stock", False)),
            "allow_closing_stock": bool(getattr(p, "allow_closing_stock", False)),
            "closing_blocked": bool(has_next_activity),

            "opening_note": getattr(p, "opening_stock_locked_note", "") or "",
            "closing_note": getattr(p, "closing_stock_locked_note", "") or "",
        })
    return JsonResponse({"ok": True, "rows": rows})

@staff_member_required
@require_POST
@transaction.atomic
def api_periods_stock_toggle(request):
    pid = request.POST.get("id")
    kind = (request.POST.get("kind") or "").strip().lower()    # opening|closing
    action = (request.POST.get("action") or "").strip().lower() # open|close

    p = Period.objects.filter(id=pid).first()
    if not p:
        return _bad("الفترة غير موجودة", 404)

    if kind not in ("opening", "closing"):
        return _bad("نوع غير صحيح")

    if action not in ("open", "close"):
        return _bad("إجراء غير صحيح")

    if kind == "opening":
        if action == "open":
            if _has_period_activity(p):
                return _bad("لا يمكن فتح جرد أول المدة لأن الفترة عليها حركة.")
            p.allow_opening_stock = True
            p.opening_stock_locked_note = ""
            p.save(update_fields=["allow_opening_stock", "opening_stock_locked_note"])
        else:
            p.allow_opening_stock = False
            p.opening_stock_locked_note = "تم قفل جرد أول المدة."
            p.save(update_fields=["allow_opening_stock", "opening_stock_locked_note"])

    if kind == "closing":
        if action == "open":
            if _has_next_period_activity(p):
                return _bad("لا يمكن فتح جرد آخر المدة لأن الفترة التالية عليها حركة.")
            p.allow_closing_stock = True
            p.closing_stock_locked_note = ""
            p.save(update_fields=["allow_closing_stock", "closing_stock_locked_note"])
        else:
            p.allow_closing_stock = False
            p.closing_stock_locked_note = "تم قفل جرد آخر المدة."
            p.save(update_fields=["allow_closing_stock", "closing_stock_locked_note"])

    return JsonResponse({"ok": True})


# =========================
# Expenses Entry APIs
# =========================
@staff_member_required
@transaction.atomic
def exp_entry_load(request):
    period_id = request.GET.get("period")
    p = Period.objects.filter(id=period_id).first() if period_id else Period.objects.order_by("-year", "-month").first()
    if not p:
        return _bad("لا توجد فترات مسجلة", 404)

    batch, _ = ExpenseBatch.objects.get_or_create(period=p)

    active_items = list(
        ExpenseItem.objects.filter(is_active=True)
        .select_related("category")
        .order_by("category__nature", "category__code", "code")
    )

    existing = {x.item_id: x for x in ExpenseLine.objects.filter(batch=batch).select_related("item", "item__category")}
    to_create = []
    for it in active_items:
        if it.id not in existing:
            to_create.append(ExpenseLine(batch=batch, item=it, amount=Decimal("0.00"), notes=""))
    if to_create:
        ExpenseLine.objects.bulk_create(to_create)
        existing = {x.item_id: x for x in ExpenseLine.objects.filter(batch=batch).select_related("item", "item__category")}

    def nature_label(n):
        return {"OP": "تشغيلي", "SA": "بيعي", "AD": "إداري"}.get(n, n)

    groups = {"OP": [], "SA": [], "AD": []}
    for it in active_items:
        line = existing.get(it.id)
        n = getattr(it.category, "nature", "OP")
        groups.setdefault(n, []).append({
            "line_id": line.id,
            "item_id": it.id,
            "item_code": it.code,
            "item_name": it.name,
            "category": str(it.category),
            "amount": str(line.amount or Decimal("0.00")),
            "notes": line.notes or "",
        })

    totals = {}
    for n in ["OP", "SA", "AD"]:
        s = ExpenseLine.objects.filter(batch=batch, item__category__nature=n).aggregate(t=Sum("amount"))["t"] or Decimal("0.00")
        totals[n] = str(s)
    totals["ALL"] = str(Decimal(totals["OP"]) + Decimal(totals["SA"]) + Decimal(totals["AD"]))

    return JsonResponse({
        "ok": True,
        "period": {"id": p.id, "label": str(p), "is_closed": p.is_closed},
        "batch_id": batch.id,
        "groups": {
            "OP": {"code": "OP", "title": f"💡 {nature_label('OP')}", "rows": groups.get("OP", [])},
            "SA": {"code": "SA", "title": f"🛒 {nature_label('SA')}", "rows": groups.get("SA", [])},
            "AD": {"code": "AD", "title": f"🏢 {nature_label('AD')}", "rows": groups.get("AD", [])},
        },
        "totals": totals,
    })

@staff_member_required
@require_http_methods(["POST"])
@transaction.atomic
def exp_entry_save(request):
    period_id = request.POST.get("period_id")
    if not period_id:
        return _bad("period_id مطلوب")

    p = Period.objects.get(id=period_id)
    if p.is_closed:
        return _bad("الفترة مقفلة، لا يمكن الحفظ", 409)

    batch, _ = ExpenseBatch.objects.get_or_create(period=p)

    line_id = request.POST.get("line_id")
    amount = _d(request.POST.get("amount"), "0")
    notes = request.POST.get("notes", "")

    if not line_id:
        return _bad("line_id مطلوب")

    line = ExpenseLine.objects.select_for_update().get(id=line_id, batch=batch)
    line.amount = amount
    line.notes = notes
    line.full_clean()
    line.save()
    return JsonResponse({"ok": True})

@staff_member_required
@require_http_methods(["POST"])
@transaction.atomic
def exp_entry_clear(request):
    period_id = request.POST.get("period_id")
    scope = request.POST.get("scope")  # group|all
    nature = request.POST.get("nature")  # OP/SA/AD

    p = Period.objects.get(id=period_id)
    if p.is_closed:
        return _bad("الفترة مقفلة، لا يمكن المسح", 409)

    batch, _ = ExpenseBatch.objects.get_or_create(period=p)

    qs = ExpenseLine.objects.filter(batch=batch)
    if scope == "group" and nature:
        qs = qs.filter(item__category__nature=nature)

    qs.update(amount=Decimal("0.00"), notes="")
    return JsonResponse({"ok": True})


# =========================
# BOM APIs (Drag & Drop)
# =========================
def _has_movements_for_bom(bom: BillOfMaterial) -> bool:
    return False

@staff_member_required
@require_GET
def bom_palette(request):
    q = (request.GET.get("q") or "").strip()

    raw_qs  = RawMaterial.objects.select_related("ingredient_unit", "storage_unit").all()
    semi_qs = Product.objects.select_related("base_unit").filter(is_semi_finished=True)

    if q:
        raw_qs  = raw_qs.filter(Q(name__icontains=q) | Q(sku__icontains=q))
        semi_qs = semi_qs.filter(Q(name__icontains=q) | Q(code__icontains=q))

    items = []

    for r in raw_qs.order_by("name")[:3000]:
        unit_id = r.ingredient_unit_id or r.storage_unit_id
        unit_obj = getattr(r, "ingredient_unit", None) or getattr(r, "storage_unit", None)
        items.append({
            "id": r.id,
            "kind": "RAW",
            "code": getattr(r, "sku", "") or "",
            "name": r.name,
            "unit_id": unit_id,
            "unit_name": str(unit_obj or ""),
            "unit_cost": str(_dec(raw_unit_cost(r))),
        })

    for p in semi_qs.order_by("code")[:3000]:
        items.append({
            "id": p.id,
            "kind": "SEMI",
            "code": p.code or "",
            "name": p.name,
            "unit_id": p.base_unit_id,
            "unit_name": str(p.base_unit) if p.base_unit_id else "",
            "unit_cost": str(_dec(semi_unit_cost(p))),
        })

    return JsonResponse({"ok": True, "items": items})

@staff_member_required
@require_GET
def bom_lock_status(request, bom_id: int):
    bom = BillOfMaterial.objects.select_related("product").get(pk=bom_id)
    return JsonResponse({"is_locked": _has_movements_for_bom(bom)})

@staff_member_required
@require_GET
def bom_get(request, bom_id: int):
    bom = BillOfMaterial.objects.select_related("product", "batch_output_unit").get(pk=bom_id)
    units = list(Unit.objects.order_by("name").values("id", "name"))

    raw_lines = []
    semi_lines = []

    qs = (
        BOMItem.objects
        .select_related(
            "raw_material",
            "raw_material__ingredient_unit",
            "raw_material__storage_unit",
            "component_product",
            "component_product__base_unit",
        )
        .filter(bom=bom)
        .order_by("id")
    )

    for it in qs:
        if it.raw_material_id:
            rm = it.raw_material
            unit_id = rm.ingredient_unit_id or rm.storage_unit_id
            unit_obj = getattr(rm, "ingredient_unit", None) or getattr(rm, "storage_unit", None)

            raw_lines.append({
                "item_id": rm.id,
                "code": rm.sku or "",
                "name": rm.name,
                "qty": str(_dec(it.quantity)),
                "unit_id": unit_id,
                "unit_name": str(unit_obj or ""),
                "unit_cost": str(_dec(raw_unit_cost(rm))),
            })

        elif it.component_product_id:
            p = it.component_product
            semi_lines.append({
                "item_id": p.id,
                "code": p.code or "",
                "name": p.name,
                "qty": str(_dec(it.quantity)),
                "unit_id": p.base_unit_id,
                "unit_name": str(p.base_unit) if p.base_unit_id else "",
                "unit_cost": str(_dec(semi_unit_cost(p))),
            })

    return JsonResponse({
        "ok": True,
        "bom": {
            "id": bom.id,
            "product_id": bom.product_id,
            "product_code": getattr(bom.product, "code", "") or "",
            "product_name": getattr(bom.product, "name", str(bom.product)),
            "name": getattr(bom, "name", "") or "",
            "is_active": bool(getattr(bom, "is_active", True)),
            "batch_output_quantity": str(getattr(bom, "batch_output_quantity", None) or ""),
            "batch_output_unit_id": getattr(bom, "batch_output_unit_id", None),
        },
        "units": units,
        "raw_lines": raw_lines,
        "semi_lines": semi_lines,
    })

@staff_member_required
@require_POST
@transaction.atomic
def bom_save(request, bom_id: int):
    bom = BillOfMaterial.objects.select_related("product").get(pk=bom_id)
    if _has_movements_for_bom(bom):
        return _bad("الوصفة مقفلة بسبب وجود حركات.")

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return _bad("بيانات غير صالحة.")

    bom.name = payload.get("name") or bom.name
    bom.is_active = bool(payload.get("is_active", bom.is_active))

    boq = payload.get("batch_output_quantity")
    bom.batch_output_quantity = _d(boq, "0") if str(boq or "").strip() != "" else None

    bou = payload.get("batch_output_unit_id")
    bom.batch_output_unit_id = int(bou) if str(bou or "").strip() != "" else None
    bom.save()

    raw_lines = payload.get("raw_lines") or []
    semi_lines = payload.get("semi_lines") or []

    BOMItem.objects.filter(bom=bom).delete()

    for line in raw_lines:
        rm = RawMaterial.objects.filter(pk=line.get("item_id")).first()
        if not rm:
            continue
        BOMItem.objects.create(
            bom=bom,
            raw_material=rm,
            component_product=None,
            quantity=_d(line.get("qty"), "0"),
        )

    for line in semi_lines:
        cp = Product.objects.filter(pk=line.get("item_id")).first()
        if not cp:
            continue
        BOMItem.objects.create(
            bom=bom,
            raw_material=None,
            component_product=cp,
            quantity=_d(line.get("qty"), "0"),
        )

    return JsonResponse({"ok": True})

@staff_member_required
@require_GET
@transaction.atomic
def bom_open_by_product(request, product_id: int):
    product = Product.objects.get(pk=product_id)
    bom = BillOfMaterial.objects.filter(product=product).order_by("-id").first()
    created = False

    if not bom:
        bom = BillOfMaterial.objects.create(
            product=product,
            name=f"وصفة {product.name}",
            is_active=True,
            batch_output_quantity=Decimal("1"),
            batch_output_unit_id=getattr(product, "base_unit_id", None),
        )
        created = True

    return JsonResponse({"bom_id": bom.id, "created": created})

@staff_member_required
@require_GET
def bom_fg_products(request):
    qs = Product.objects.filter(is_semi_finished=False).order_by("code")
    items = [{"id": p.id, "code": p.code, "name": p.name} for p in qs[:4000]]
    return JsonResponse({"items": items})


# =========================
# Inventory (StockCount)
# =========================
def inv_has_movements_for_count(count: StockCount) -> bool:
    return False

def inv_is_locked(count: StockCount) -> bool:
    if getattr(count.period, "is_closed", False):
        return True
    if getattr(count, "is_committed", False):
        return True
    if inv_has_movements_for_count(count):
        return True
    return False

def inv_next_period_has_movements(period: Period) -> bool:
    return False

def _ensure_count(period: Period, count_type: str) -> StockCount:
    count, _ = StockCount.objects.get_or_create(
        period=period,
        type=count_type,
        defaults={
            "count_type": count_type,
            "count_date": getattr(period, "start_date", None) or getattr(period, "end_date", None),
            "notes": "",
        },
    )
    if getattr(count, "count_type", None) != count_type:
        count.count_type = count_type
        count.save(update_fields=["count_type"])
    return count

def _prefill_stockcount_lines(count: StockCount) -> int:
    if inv_is_locked(count):
        return 0

    existing_raw = set(
        count.lines.filter(raw_material_id__isnull=False).values_list("raw_material_id", flat=True)
    )
    existing_semi = set(
        count.lines.filter(semi_finished_product_id__isnull=False).values_list("semi_finished_product_id", flat=True)
    )

    to_create = []

    for rid, storage_unit_id, ingredient_unit_id in RawMaterial.objects.values_list("id", "storage_unit_id", "ingredient_unit_id"):
        if rid in existing_raw:
            continue
        default_unit = storage_unit_id or ingredient_unit_id
        if not default_unit:
            continue
        to_create.append(
            StockCountLine(
                stock_count=count,
                raw_material_id=rid,
                semi_finished_product_id=None,
                unit_id=default_unit,
                quantity=Decimal("0"),
            )
        )

    for pid, base_unit_id in Product.objects.filter(is_semi_finished=True).values_list("id", "base_unit_id"):
        if pid in existing_semi:
            continue
        if not base_unit_id:
            continue
        to_create.append(
            StockCountLine(
                stock_count=count,
                raw_material_id=None,
                semi_finished_product_id=pid,
                unit_id=base_unit_id,
                quantity=Decimal("0"),
            )
        )

    if to_create:
        StockCountLine.objects.bulk_create(to_create, batch_size=1000)

    return len(to_create)

@staff_member_required
@require_GET
def inv_stockcount_state(request):
    period_id = request.GET.get("period_id") or ""
    req_type = (request.GET.get("type") or "opening").strip().lower()
    if req_type not in ("opening", "closing"):
        req_type = "opening"

    period = get_object_or_404(Period, id=period_id)

    if req_type == "opening":
        if not getattr(period, "inv_opening_enabled", False):
            return JsonResponse({"ok": False, "error": "⛔ جرد أول الفترة مقفول. افتحه أولاً من شاشة الفترات."}, status=403)

    count = _ensure_count(period, req_type)
    locked = inv_is_locked(count)

    _prefill_stockcount_lines(count)

    lines = []
    for ln in count.lines.select_related("raw_material", "semi_finished_product", "unit").order_by("id"):
        item = ln.raw_material or ln.semi_finished_product
        item_type = "raw" if ln.raw_material_id else "semi"

        if count.type == "opening":
            uc = ln.unit_cost_value
            tc = (uc * (ln.quantity or Decimal("0"))) if uc is not None else None
            unit_cost = uc
            total_cost = tc
        else:
            unit_cost = ln.saved_unit_cost if ln.saved_unit_cost is not None else ln.unit_cost()
            total_cost = ln.saved_total_cost if ln.saved_total_cost is not None else ln.line_total_cost()

        lines.append({
            "id": ln.id,
            "item_type": item_type,
            "item_id": item.id if item else None,
            "item_label": str(item) if item else "",
            "unit_id": ln.unit_id,
            "unit_label": str(ln.unit) if ln.unit_id else "",
            "qty": _s(ln.quantity),
            "unit_cost": _s(unit_cost),
            "total_cost": _s(total_cost),
        })

    return JsonResponse({
        "ok": True,
        "requested_type": req_type,
        "effective_type": req_type,
        "locked": locked,
        "count_id": count.id,
        "period": {"id": period.id, "label": str(period), "is_closed": bool(getattr(period, "is_closed", False))},
        "count_type": req_type,
        "totals": {"total_qty": _s(count.total_quantity()), "total_cost": _s(count.total_cost())},
        "lines": lines,
    })

@staff_member_required
@require_POST
@transaction.atomic
def inv_stockcount_update_line(request):
    line_id = request.POST.get("line_id")
    ln = get_object_or_404(StockCountLine, id=line_id)
    count = ln.stock_count

    if inv_is_locked(count):
        return _bad("❌ ممنوع التعديل: الفترة مقفولة أو يوجد حركة.", 400)

    unit_id = request.POST.get("unit_id")
    qty = request.POST.get("qty")
    unit_cost_value = request.POST.get("unit_cost_value", None)

    if unit_id:
        ln.unit_id = int(unit_id)

    if qty is not None:
        q = _d(qty)
        if q < 0:
            return _bad("الكمية لا يمكن أن تكون سالبة.", 400)
        ln.quantity = q

    if unit_cost_value is not None:
        if count.type != "opening":
            return _bad("❌ تكلفة الوحدة اليدوية مسموحة في جرد أول المدة فقط.", 400)
        v = str(unit_cost_value).strip()
        if v == "":
            ln.unit_cost_value = None
        else:
            c = _d(v)
            if c < 0:
                return _bad("التكلفة لا يمكن أن تكون سالبة.", 400)
            ln.unit_cost_value = c

    ln.save()
    return JsonResponse({"ok": True})

@staff_member_required
@require_POST
@transaction.atomic
def inv_stockcount_clear_all(request):
    count_id = request.POST.get("count_id")
    count = get_object_or_404(StockCount, id=count_id)

    if inv_is_locked(count):
        return _bad("❌ ممنوع: الفترة مقفولة أو يوجد حركة.", 400)

    _prefill_stockcount_lines(count)

    qs = count.lines.all()
    qs.update(quantity=Decimal("0"))
    if count.type == "opening":
        qs.update(unit_cost_value=None)

    for ln in count.lines.all():
        ln.save()

    return JsonResponse({"ok": True})

@staff_member_required
@require_POST
@transaction.atomic
def inv_stockcount_commit(request):
    count_id = request.POST.get("count_id")
    count = get_object_or_404(StockCount, id=count_id)

    if inv_is_locked(count):
        return _bad("❌ الجرد مقفول ولا يمكن اعتماده.", 400)

    if count.type == "opening":
        if not getattr(count.period, "inv_opening_enabled", False):
            return _bad("⛔ لا يمكن اعتماد جرد أول الفترة قبل فتحه من Period.", 400)

    count.is_committed = True
    count.committed_at = timezone.now()
    count.save(update_fields=["is_committed", "committed_at"])
    return JsonResponse({"ok": True})


# =========================
# Sales Grid APIs
# =========================
@staff_member_required
@require_GET
def api_sales_grid_get(request):
    period_id = request.GET.get("period_id")
    if not period_id:
        return _bad("period_id مطلوب")

    period = Period.objects.filter(id=period_id).first()
    if not period:
        return _bad("الفترة غير موجودة")

    summary, _ = SalesSummary.objects.get_or_create(period=period)

    products = (
        Product.objects
        .filter(is_sellable=True, is_semi_finished=False)
        .select_related("base_unit")
        .order_by("code", "name")
    )

    existing = {l.product_id: l for l in summary.lines.select_related("product", "unit").all()}

    rows = []
    for p in products:
        l = existing.get(p.id)
        default_unit_id = p.base_unit_id

        rows.append({
            "line_id": l.id if l else None,
            "product_id": p.id,
            "product_name": p.name,
            "code": p.code or "",
            "unit_id": (l.unit_id if l else default_unit_id),
            "unit_name": (str(l.unit) if l and l.unit_id else (str(p.base_unit) if p.base_unit_id else "")),
            "quantity": str(l.quantity or 0) if l else "0",
            "unit_price": str(l.unit_price or 0) if l else str(getattr(p, "selling_price_per_unit", 0) or "0"),
            "line_total": str(l.line_total or 0) if l else "0",
        })

    return JsonResponse({
        "ok": True,
        "is_locked": is_period_locked(period),
        "summary_id": summary.id,
        "rows": rows
    })

@staff_member_required
@require_POST
@transaction.atomic
def api_sales_grid_save(request):
    period_id = request.POST.get("period_id")
    rows_json = request.POST.get("rows_json")

    if not period_id:
        return _bad("period_id مطلوب")
    if not rows_json:
        return _bad("rows_json مطلوب")

    period = Period.objects.filter(id=period_id).first()
    if not period:
        return _bad("الفترة غير موجودة")

    if is_period_locked(period):
        return _bad("الفترة مقفولة - ممنوع الحفظ")

    try:
        rows = json.loads(rows_json)
        if not isinstance(rows, list):
            return _bad("rows_json لازم يكون List")
    except Exception:
        return _bad("rows_json غير صالح")

    summary, _ = SalesSummary.objects.get_or_create(period=period)

    existing = {l.product_id: l for l in SalesSummaryLine.objects.select_for_update().filter(summary=summary)}
    saved = 0

    for r in rows:
        product_id = r.get("product_id")
        unit_id = r.get("unit_id")
        qty = _d(r.get("quantity"))
        price = _d(r.get("unit_price"))

        if not product_id:
            continue

        if not unit_id:
            unit_id = Product.objects.filter(id=product_id).values_list("base_unit_id", flat=True).first()
        if not unit_id:
            continue

        if qty <= 0:
            old = existing.get(int(product_id))
            if old:
                old.delete()
            continue

        product = Product.objects.filter(id=product_id).first()
        unit = Unit.objects.filter(id=unit_id).first()
        if not product or not unit:
            continue

        line = existing.get(int(product_id))
        if not line:
            line = SalesSummaryLine(summary=summary, product=product)

        line.unit = unit
        line.quantity = qty
        line.unit_price = price
        line.save()
        saved += 1

    summary.save()
    return JsonResponse({"ok": True, "saved_count": saved})


# =========================
# Inventory Palette (Legacy)
# =========================
@staff_member_required
@require_GET
def inv_stockcount_palette(request):
    """
    ✅ Legacy endpoint
    لو شاشتك الجديدة مش بتستخدمه، نخليه موجود عشان urls.py ما يقعش.
    يرجع قائمة RAW و SEMI بشكل بسيط.
    """
    q = (request.GET.get("q") or "").strip()

    raw_qs = RawMaterial.objects.all()
    semi_qs = Product.objects.filter(is_semi_finished=True)

    if q:
        raw_qs = raw_qs.filter(Q(name__icontains=q) | Q(sku__icontains=q))
        semi_qs = semi_qs.filter(Q(name__icontains=q) | Q(code__icontains=q))

    raw = [
        {
            "type": "raw",
            "id": r.id,
            "label": f"{(r.sku or '').strip()} - {r.name}".strip(" -"),
            "default_unit_id": (r.storage_unit_id or r.ingredient_unit_id),
        }
        for r in raw_qs.order_by("name")[:4000]
    ]

    semi = [
        {
            "type": "semi",
            "id": p.id,
            "label": f"{p.code} - {p.name}",
            "default_unit_id": p.base_unit_id,
        }
        for p in semi_qs.order_by("code")[:4000]
    ]

    return JsonResponse({"ok": True, "raw": raw, "semi": semi})

@staff_member_required
@require_GET
def inv_units_list(request):
    q = (request.GET.get("q") or "").strip()
    qs = Unit.objects.all()
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(abbreviation__icontains=q))
    data = [{"id": u.id, "label": str(u)} for u in qs.order_by("name")[:250]]
    return JsonResponse({"ok": True, "units": data})



@staff_member_required
@require_POST
@transaction.atomic
def inv_stockcount_add_line(request):
    count_id = request.POST.get("count_id")
    count = get_object_or_404(StockCount, id=count_id)

    if inv_is_locked(count):
        return JsonResponse({"ok": False, "error": "❌ ممنوع الحفظ/الإضافة: الفترة مقفولة أو يوجد حركة."}, status=400)

    item_type = (request.POST.get("item_type") or "").strip()
    item_id = request.POST.get("item_id") or ""
    unit_id = request.POST.get("unit_id") or ""
    qty = _d(request.POST.get("qty"))

    if item_type not in ("raw", "semi"):
        return JsonResponse({"ok": False, "error": "نوع الصنف غير صحيح."}, status=400)
    if not item_id or not unit_id:
        return JsonResponse({"ok": False, "error": "اختر الصنف والوحدة."}, status=400)
    if qty < 0:
        return JsonResponse({"ok": False, "error": "الكمية لا يمكن أن تكون سالبة."}, status=400)

    ln = StockCountLine(stock_count=count, unit_id=int(unit_id), quantity=qty)
    if item_type == "raw":
        ln.raw_material_id = int(item_id)
        ln.semi_finished_product_id = None
    else:
        ln.semi_finished_product_id = int(item_id)
        ln.raw_material_id = None

    ln.save()
    return JsonResponse({"ok": True, "id": ln.id})


@staff_member_required
@require_POST
@transaction.atomic
def inv_stockcount_update_line(request):
    line_id = request.POST.get("line_id")
    ln = get_object_or_404(StockCountLine, id=line_id)
    count = ln.stock_count

    if inv_is_locked(count):
        return JsonResponse({"ok": False, "error": "❌ ممنوع التعديل: الفترة مقفولة أو يوجد حركة."}, status=400)

    unit_id = request.POST.get("unit_id")
    qty = request.POST.get("qty")
    unit_cost_value = request.POST.get("unit_cost_value", None)

    if unit_id:
        ln.unit_id = int(unit_id)

    if qty is not None:
        q = _d(qty)
        if q < 0:
            return JsonResponse({"ok": False, "error": "الكمية لا يمكن أن تكون سالبة."}, status=400)
        ln.quantity = q

    if unit_cost_value is not None:
        if count.type != "opening":
            return JsonResponse({"ok": False, "error": "❌ تكلفة الوحدة اليدوية مسموحة في جرد أول المدة فقط."}, status=400)

        v = str(unit_cost_value).strip()
        if v == "":
            ln.unit_cost_value = None
        else:
            c = _d(v)
            if c < 0:
                return JsonResponse({"ok": False, "error": "التكلفة لا يمكن أن تكون سالبة."}, status=400)
            ln.unit_cost_value = c

    ln.save()
    return JsonResponse({"ok": True})


@staff_member_required
@require_POST
@transaction.atomic
def inv_stockcount_delete_line(request):
    line_id = request.POST.get("line_id")
    ln = get_object_or_404(StockCountLine, id=line_id)

    if inv_is_locked(ln.stock_count):
        return JsonResponse({"ok": False, "error": "❌ ممنوع: الفترة مقفولة أو يوجد حركة."}, status=400)

    ln.quantity = Decimal("0")
    if ln.stock_count.type == "opening":
        ln.unit_cost_value = None
    ln.save()
    return JsonResponse({"ok": True})


@staff_member_required
@require_POST
@transaction.atomic
def inv_stockcount_recalc(request):
    count_id = request.POST.get("count_id")
    count = get_object_or_404(StockCount, id=count_id)

    if inv_is_locked(count):
        return JsonResponse({"ok": False, "error": "❌ ممنوع: الفترة مقفولة أو يوجد حركة."}, status=400)

    _prefill_stockcount_lines(count)
    for ln in count.lines.all():
        ln.save()

    return JsonResponse({"ok": True})


@staff_member_required
@require_POST
@transaction.atomic
def inv_stockcount_clear_all(request):
    count_id = request.POST.get("count_id")
    count = get_object_or_404(StockCount, id=count_id)

    if inv_is_locked(count):
        return JsonResponse({"ok": False, "error": "❌ ممنوع: الفترة مقفولة أو يوجد حركة."}, status=400)

    _prefill_stockcount_lines(count)

    qs = count.lines.all()
    qs.update(quantity=Decimal("0"))

    if count.type == "opening":
        qs.update(unit_cost_value=None)

    for ln in count.lines.all():
        ln.save()

    return JsonResponse({"ok": True})


@staff_member_required
@require_POST
@transaction.atomic
def inv_stockcount_submit(request):
    count_id = request.POST.get("count_id")
    count = get_object_or_404(StockCount, id=count_id)

    if inv_is_locked(count):
        return JsonResponse({"ok": False, "error": "❌ الجرد مقفول ولا يمكن حفظه."}, status=400)

    if count.type == "closing":
        if inv_next_period_has_movements(count.period):
            return JsonResponse({"ok": False, "error": "❌ لا يمكن إقفال جرد آخر الفترة: توجد حركات في الفترة التالية."}, status=400)

    count.is_submitted = True
    count.save(update_fields=["is_submitted"])
    return JsonResponse({"ok": True})


@staff_member_required
@require_POST
@transaction.atomic
def inv_stockcount_commit(request):
    count_id = request.POST.get("count_id")
    count = get_object_or_404(StockCount, id=count_id)

    if inv_is_locked(count):
        return JsonResponse({"ok": False, "error": "❌ الجرد مقفول ولا يمكن اعتماده."}, status=400)

    if count.type == "opening":
        if not getattr(count.period, "inv_opening_enabled", False):
            return JsonResponse({"ok": False, "error": "⛔ لا يمكن اعتماد جرد أول الفترة قبل فتحه من Period."}, status=400)

    count.is_committed = True
    count.committed_at = timezone.now()
    count.save(update_fields=["is_committed", "committed_at"])
    return JsonResponse({"ok": True})

# portal/api.py (add this section)
import json
from decimal import Decimal
from django.http import JsonResponse
from django.views.decorators.http import require_GET, require_POST
from django.contrib.admin.views.decorators import staff_member_required
from django.db import transaction

from expenses.models import Period
from costing.models import Product, Unit
from sales.models import SalesSummary, SalesSummaryLine


def _d(v, default="0"):
    try:
        return Decimal(str(v if v is not None else default))
    except Exception:
        return Decimal(default)


def is_period_locked(period: Period) -> bool:
    return bool(getattr(period, "is_closed", False))


@staff_member_required
@require_GET
def portal_sales_grid_get(request):
    period_id = request.GET.get("period_id")
    if not period_id:
        return JsonResponse({"ok": False, "error": "period_id مطلوب"}, status=400)

    period = Period.objects.filter(id=period_id).first()
    if not period:
        return JsonResponse({"ok": False, "error": "الفترة غير موجودة"}, status=404)

    summary, _ = SalesSummary.objects.get_or_create(period=period)

    # منتجات بيع (FG) فقط
    products = (
        Product.objects
        .filter(is_semi_finished=False)
        .select_related("base_unit")
        .order_by("code", "name")
    )

    # لاحظ: unit هنا تابع لـ SalesSummaryLine (مسموح)
    existing = {
        l.product_id: l
        for l in summary.lines.select_related("product", "unit").all()
    }

    rows = []
    for p in products:
        l = existing.get(p.id)

        default_unit_id = getattr(p, "base_unit_id", None)
        default_unit_name = str(p.base_unit) if getattr(p, "base_unit_id", None) else ""

        rows.append({
            "line_id": l.id if l else None,
            "product_id": p.id,
            "code": getattr(p, "code", "") or "",
            "name": str(p),
            "unit_id": (l.unit_id if l else default_unit_id),
            "unit_name": (str(l.unit) if l and l.unit_id else default_unit_name),
            "quantity": str(l.quantity or 0) if l else "0",
            "unit_price": str(l.unit_price or 0) if l else "0",
            "line_total": str(l.line_total or 0) if l else "0",
        })

    return JsonResponse({
        "ok": True,
        "period": {"id": period.id, "label": str(period), "is_closed": bool(period.is_closed)},
        "is_locked": is_period_locked(period),
        "summary_id": summary.id,
        "rows": rows,
    })

@staff_member_required
@require_POST
@transaction.atomic
def portal_sales_grid_save(request):
    period_id = request.POST.get("period_id")
    rows_json = request.POST.get("rows_json")

    if not period_id:
        return JsonResponse({"ok": False, "error": "period_id مطلوب"}, status=400)
    if not rows_json:
        return JsonResponse({"ok": False, "error": "rows_json مطلوب"}, status=400)

    period = Period.objects.filter(id=period_id).first()
    if not period:
        return JsonResponse({"ok": False, "error": "الفترة غير موجودة"}, status=404)
    if is_period_locked(period):
        return JsonResponse({"ok": False, "error": "الفترة مقفلة - ممنوع الحفظ"}, status=409)

    try:
        rows = json.loads(rows_json)
        if not isinstance(rows, list):
            return JsonResponse({"ok": False, "error": "rows_json لازم يكون List"}, status=400)
    except Exception:
        return JsonResponse({"ok": False, "error": "rows_json غير صالح"}, status=400)

    summary, _ = SalesSummary.objects.get_or_create(period=period)

    existing = {
        l.product_id: l
        for l in SalesSummaryLine.objects.select_for_update().filter(summary=summary)
    }

    saved = 0
    for r in rows:
        product_id = r.get("product_id")
        unit_id = r.get("unit_id")
        qty = _d(r.get("quantity"))
        price = _d(r.get("unit_price"))

        if not product_id or not unit_id:
            continue

        product = Product.objects.filter(id=product_id).first()
        unit = Unit.objects.filter(id=unit_id).first()
        if not product or not unit:
            continue

        # لو الكمية صفر -> احذف السطر إن وجد
        if qty <= 0:
            l = existing.get(int(product_id))
            if l:
                l.delete()
            continue

        line = existing.get(int(product_id))
        if not line:
            line = SalesSummaryLine(summary=summary, product=product)

        line.unit = unit
        line.quantity = qty
        line.unit_price = price
        line.save()  # يحسب line_total من الموديل
        saved += 1

    summary.save()
    return JsonResponse({"ok": True, "saved_count": saved})

################################################################################################################################
# portal/api.py
@staff_member_required
@require_GET
@transaction.atomic
def purchases_grid(request):
    period_id = request.GET.get("period_id")
    period = Period.objects.get(id=period_id)

    summary, _ = PurchaseSummary.objects.get_or_create(period=period)

    rows = []

    # 1️⃣ خامات
    for rm in RawMaterial.objects.order_by("name"):
        line, _ = PurchaseSummaryLine.objects.get_or_create(
            summary=summary,
            raw_material=rm,
            defaults={
                "unit": rm.ingredient_unit or rm.storage_unit,
                "quantity": 0,
                "unit_price": rm.purchase_price_per_storage_unit or 0,
            },
        )

        rows.append({
            "line_id": line.id,
            "item_type": "RAW",
            "code": rm.sku,
            "name": rm.name,
            "unit_id": line.unit_id,
            "unit_name": str(line.unit),
            "quantity": str(line.quantity),
            "unit_price": str(line.unit_price),
        })

    # 2️⃣ نصف مصنع (اختياري)
    for p in Product.objects.filter(is_semi_finished=True):
        line, _ = PurchaseSummaryLine.objects.get_or_create(
            summary=summary,
            semi_product=p,
            defaults={
                "unit": p.base_unit,
                "quantity": 0,
                "unit_price": p.compute_unit_cost(period),
            },
        )

        rows.append({
            "line_id": line.id,
            "item_type": "SEMI",
            "code": p.code,
            "name": p.name,
            "unit_id": line.unit_id,
            "unit_name": str(line.unit),
            "quantity": str(line.quantity),
            "unit_price": str(line.unit_price),
        })

    return JsonResponse({
        "ok": True,
        "rows": rows,
        "is_locked": period.is_closed,
    })

##############################################################################################################

@staff_member_required
@require_POST
@transaction.atomic
def purchases_grid_save(request):
    period = Period.objects.get(id=request.POST.get("period_id"))

    if period.is_closed:
        return JsonResponse({"ok": False, "error": "الفترة مقفلة"}, status=409)

    rows = json.loads(request.POST.get("rows_json") or "[]")
    saved = 0

    for r in rows:
        line = PurchaseSummaryLine.objects.select_for_update().get(id=r["line_id"])
        line.quantity = Decimal(r["quantity"])
        line.unit_price = Decimal(r["unit_price"])
        line.save()
        saved += 1

    return JsonResponse({"ok": True, "saved_count": saved})

# =========================
# Purchases Grid APIs
# =========================
# portal/api.py
from django.http import JsonResponse
from django.views.decorators.http import require_GET, require_POST
from django.contrib.admin.views.decorators import staff_member_required
from django.db import transaction
from decimal import Decimal
import json

from expenses.models import Period
from costing.models import RawMaterial, Unit
from django.apps import apps

def _d(x, default="0"):
    try:
        return Decimal(str(x).replace(",", "").strip() or default)
    except Exception:
        return Decimal(default)

def _pick_cost_from_raw(rm):
    """
    ضع هنا أسماء الحقول الموجودة فعليًا في RawMaterial عندك.
    سيأخذ أول قيمة غير صفرية.
    """
    candidates = [
        "purchase_price_per_storage_unit",
        "purchase_price",
        "unit_cost",
        "standard_cost",
        "avg_cost",
        "last_cost",
        "cost",
    ]
    for fn in candidates:
        if hasattr(rm, fn):
            v = getattr(rm, fn)
            if v not in (None, "") and Decimal(str(v or 0)) != 0:
                return Decimal(str(v))
    return Decimal("0")

@staff_member_required
@require_GET
def api_purchases_grid_get(request):
    period_id = request.GET.get("period_id")
    if not period_id:
        return JsonResponse({"ok": False, "error": "period_id مطلوب"}, status=400)

    period = Period.objects.filter(id=period_id).first()
    if not period:
        return JsonResponse({"ok": False, "error": "الفترة غير موجودة"}, status=404)

    PurchaseSummary = apps.get_model("purchases", "PurchaseSummary")
    PurchaseSummaryLine = apps.get_model("purchases", "PurchaseSummaryLine")

    summary, _ = PurchaseSummary.objects.get_or_create(period=period)

    # ✅ موديلك: purchase_unit وليس unit
    existing = {
        l.raw_material_id: l
        for l in PurchaseSummaryLine.objects.filter(summary=summary)
            .select_related("raw_material", "purchase_unit")
    }

    rms = RawMaterial.objects.select_related("storage_unit").order_by("sku", "name")

    rows = []
    for rm in rms:
        line = existing.get(rm.id)

        # افتراض وحدة الشراء = storage_unit (عدّلها لو عندك purchase_unit داخل الخام)
        default_unit_id = getattr(rm, "storage_unit_id", None)
        default_unit_name = str(getattr(rm, "storage_unit", "") or "")

        default_cost = _pick_cost_from_raw(rm)  # ✅ هنا الحل: تكلفة افتراضية من الخام

        rows.append({
            "line_id": line.id if line else None,
            "raw_material_id": rm.id,
            "code": getattr(rm, "sku", "") or "",
            "name": getattr(rm, "name", "") or str(rm),

            "purchase_unit_id": (line.purchase_unit_id if line else default_unit_id),
            "unit_name": (str(line.purchase_unit) if line and line.purchase_unit_id else default_unit_name),

            "quantity": str(line.quantity if line else Decimal("0")),
            "unit_cost": str(line.unit_cost if line else default_cost),
            "line_total": str(line.line_total if line else Decimal("0")),
        })

    return JsonResponse({
        "ok": True,
        "is_locked": getattr(period, "is_closed", False),
        "rows": rows,
        "summary_total": str(summary.total_amount or 0),
    })




@staff_member_required
@require_POST
@transaction.atomic
def api_purchases_grid_save(request):
    period_id = request.POST.get("period_id")
    rows_json = request.POST.get("rows_json")

    if not period_id:
        return JsonResponse({"ok": False, "error": "period_id مطلوب"}, status=400)
    if not rows_json:
        return JsonResponse({"ok": False, "error": "rows_json مطلوب"}, status=400)

    period = Period.objects.filter(id=period_id).first()
    if not period:
        return JsonResponse({"ok": False, "error": "الفترة غير موجودة"}, status=404)

    if getattr(period, "is_closed", False):
        return JsonResponse({"ok": False, "error": "الفترة مقفلة - ممنوع الحفظ"}, status=409)

    try:
        rows = json.loads(rows_json)
        if not isinstance(rows, list):
            return JsonResponse({"ok": False, "error": "rows_json لازم يكون List"}, status=400)
    except Exception:
        return JsonResponse({"ok": False, "error": "rows_json غير صالح"}, status=400)

    PurchaseSummary = apps.get_model("purchases", "PurchaseSummary")
    PurchaseSummaryLine = apps.get_model("purchases", "PurchaseSummaryLine")

    summary, _ = PurchaseSummary.objects.get_or_create(period=period)

    existing = {
        l.raw_material_id: l
        for l in PurchaseSummaryLine.objects.select_for_update().filter(summary=summary)
    }

    saved = 0

    for r in rows:
        raw_material_id = r.get("raw_material_id")
        purchase_unit_id = r.get("purchase_unit_id")
        qty = _d(r.get("quantity"))
        cost = _d(r.get("unit_cost"))

        if not raw_material_id or not purchase_unit_id:
            continue

        rm = RawMaterial.objects.filter(id=raw_material_id).first()
        u = Unit.objects.filter(id=purchase_unit_id).first()
        if not rm or not u:
            continue

        # لو تريد حذف السطر عند qty=0:
        if qty <= 0:
            old = existing.get(int(raw_material_id))
            if old:
                old.delete()
            continue

        line = existing.get(int(raw_material_id))
        if not line:
            line = PurchaseSummaryLine(summary=summary, raw_material=rm)

        line.purchase_unit = u
        line.quantity = qty
        line.unit_cost = cost
        line.save()  # سيحسب line_total ويعيد حساب summary تلقائيًا من موديلك
        saved += 1

    return JsonResponse({
        "ok": True,
        "saved_count": saved,
        "summary_total": str(summary.total_amount or 0),
    })


# portal/api.py
from io import BytesIO
from decimal import Decimal
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET, require_POST
from django.contrib.admin.views.decorators import staff_member_required
from django.apps import apps
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

@staff_member_required
@require_GET
def api_purchases_export_xlsx(request):
    period_id = request.GET.get("period_id")
    if not period_id:
        return JsonResponse({"ok": False, "error": "period_id مطلوب"}, status=400)

    Period = apps.get_model("expenses", "Period")
    RawMaterial = apps.get_model("costing", "RawMaterial")
    PurchaseSummary = apps.get_model("purchases", "PurchaseSummary")
    PurchaseSummaryLine = apps.get_model("purchases", "PurchaseSummaryLine")

    period = Period.objects.filter(id=period_id).first()
    if not period:
        return JsonResponse({"ok": False, "error": "الفترة غير موجودة"}, status=404)

    summary, _ = PurchaseSummary.objects.get_or_create(period=period)

    existing = {
        l.raw_material_id: l
        for l in PurchaseSummaryLine.objects.filter(summary=summary)
            .select_related("raw_material", "purchase_unit")
    }

    rms = RawMaterial.objects.select_related("storage_unit").order_by("sku", "name")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchases"

    # Header
    headers = [
        "raw_material_id", "purchase_unit_id",
        "الكود", "المادة الخام", "وحدة الشراء",
        "الكمية", "تكلفة الوحدة", "الإجمالي"
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    # widths
    widths = [16, 16, 14, 38, 18, 16, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    # rows
    for rm in rms:
        line = existing.get(rm.id)

        default_unit_id = getattr(rm, "storage_unit_id", None)
        default_unit_name = str(getattr(rm, "storage_unit", "") or "")

        # لو ما فيه line محفوظ: خلّيها 0، المستخدم يعبّيها في الإكسل
        qty = (line.quantity if line else Decimal("0"))
        cost = (line.unit_cost if line else Decimal("0"))
        total = (line.line_total if line else Decimal("0"))

        pu_id = (line.purchase_unit_id if line else default_unit_id)
        pu_name = (str(line.purchase_unit) if line and line.purchase_unit_id else default_unit_name)

        ws.append([
            rm.id,
            pu_id or "",
            getattr(rm, "sku", "") or "",
            getattr(rm, "name", "") or str(rm),
            pu_name,
            float(qty),
            float(cost),
            float(total),
        ])

    # Number formats: qty 4, cost 6, total 2
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 6).number_format = '#,##0.0000'
        ws.cell(r, 7).number_format = '#,##0.000000'
        ws.cell(r, 8).number_format = '#,##0.00'

        # alignment
        for col in [1,2,3,5,6,7,8]:
            ws.cell(r, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, 4).alignment = Alignment(horizontal="right", vertical="center")

    # Hide technical columns (IDs) but keep them for import
    ws.column_dimensions["A"].hidden = True
    ws.column_dimensions["B"].hidden = True

    # Footer total (optional)
    last = ws.max_row + 1
    ws.cell(last, 5).value = "الإجمالي"
    ws.cell(last, 5).font = Font(bold=True)
    ws.cell(last, 8).value = f"=SUM(H2:H{ws.max_row})"
    ws.cell(last, 8).number_format = '#,##0.00'
    ws.cell(last, 8).font = Font(bold=True)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"purchases_{period.year}_{int(period.month):02d}.xlsx"
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# portal/api.py
from io import BytesIO
from decimal import Decimal, InvalidOperation
import openpyxl
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.admin.views.decorators import staff_member_required
from django.apps import apps

def _dec(x, default="0"):
    try:
        if x is None:
            return Decimal(default)
        s = str(x).replace(",", "").strip()
        if s == "":
            return Decimal(default)
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal(default)

@staff_member_required
@require_POST
def api_purchases_import_xlsx(request):
    f = request.FILES.get("file")
    if not f:
        return JsonResponse({"ok": False, "error": "file مطلوب"}, status=400)

    try:
        wb = openpyxl.load_workbook(filename=BytesIO(f.read()), data_only=True)
    except Exception:
        return JsonResponse({"ok": False, "error": "ملف Excel غير صالح"}, status=400)

    ws = wb.active  # أو wb["Purchases"] لو تحب تقييد الاسم

    # expected headers in row 1
    header = [str(ws.cell(1, c).value or "").strip() for c in range(1, 9)]
    expected = ["raw_material_id","purchase_unit_id","الكود","المادة الخام","وحدة الشراء","الكمية","تكلفة الوحدة","الإجمالي"]
    if header[:8] != expected:
        return JsonResponse({
            "ok": False,
            "error": "تنسيق الملف غير مطابق. استخدم ملف التصدير كقالب (نفس الأعمدة والترتيب)."
        }, status=400)

    RawMaterial = apps.get_model("costing", "RawMaterial")
    Unit = apps.get_model("costing", "Unit")

    rows = []
    count = 0

    for r in range(2, ws.max_row + 1):
        raw_id = ws.cell(r, 1).value
        unit_id = ws.cell(r, 2).value

        if raw_id in (None, ""):
            continue

        try:
            raw_id = int(raw_id)
        except Exception:
            continue

        rm = RawMaterial.objects.filter(id=raw_id).first()
        if not rm:
            continue

        # unit id optional: لو فاضي، خذ storage_unit
        if unit_id in (None, ""):
            unit_id = getattr(rm, "storage_unit_id", None)

        try:
            unit_id = int(unit_id) if unit_id else None
        except Exception:
            unit_id = None

        u = Unit.objects.filter(id=unit_id).first() if unit_id else None
        if not u:
            # حاول fallback إلى storage_unit لو موجود
            su_id = getattr(rm, "storage_unit_id", None)
            u = Unit.objects.filter(id=su_id).first() if su_id else None

        if not u:
            continue

        qty = _dec(ws.cell(r, 6).value, "0")
        cost = _dec(ws.cell(r, 7).value, "0")

        # لو كله صفر تجاهله (اختياري)
        if qty == 0 and cost == 0:
            continue

        rows.append({
            "line_id": None,
            "raw_material_id": rm.id,
            "code": getattr(rm, "sku", "") or "",
            "name": getattr(rm, "name", "") or str(rm),
            "purchase_unit_id": u.id,
            "unit_name": str(u),
            "quantity": str(qty),     # DB: 4 decimals
            "unit_cost": str(cost),   # DB: 6 decimals
            "line_total": "0",
        })
        count += 1

    return JsonResponse({"ok": True, "rows": rows, "count": count})



from decimal import Decimal, InvalidOperation

def _dec(x, default="0"):
    try:
        if x is None:
            return Decimal(default)
        s = str(x).replace(",", "").strip()
        if s == "":
            return Decimal(default)
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal(default)






from io import BytesIO
from decimal import Decimal, InvalidOperation
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET
from django.contrib.admin.views.decorators import staff_member_required
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import json

def _dec(x, default="0"):
    try:
        if x is None:
            return Decimal(default)
        s = str(x).replace(",", "").strip()
        if s == "":
            return Decimal(default)
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal(default)

@staff_member_required
@require_GET
def api_sales_export_xlsx(request):
    period_id = request.GET.get("period_id")
    if not period_id:
        return JsonResponse({"ok": False, "error": "period_id مطلوب"}, status=400)

    # ✅ استخدم grid_get (زي ما عندك في api.py) لأنه بيرجع نفس الصفوف المعروضة في الشاشة
    grid_resp = api_sales_grid_get(request)
    data = json.loads(grid_resp.content.decode("utf-8"))

    if not data.get("ok"):
        return JsonResponse({"ok": False, "error": data.get("error", "خطأ")}, status=400)

    rows = data.get("rows") or []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    # ✅ (اختياري) عمود رقم تسلسلي
    headers = [
        "#",
        "product_id", "unit_id",
        "الكود", "المنتج", "الوحدة",
        "الكمية", "سعر الوحدة", "الإجمالي"
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="0D6EFD")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    widths = [6, 12, 10, 14, 40, 14, 14, 16, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    # ✅ مهم: لا تستخدم rows[1:] ولا أي شرط يطرد أول صف
    for idx, r in enumerate(rows, start=1):
        qty   = _dec(r.get("quantity"), "0")
        price = _dec(r.get("unit_price"), "0")

        # ✅ الإجمالي: لو موجود line_total استخدمه، وإلا احسبه
        if r.get("line_total") not in (None, ""):
            total = _dec(r.get("line_total"), "0")
        else:
            total = qty * price

        # ✅ اسم المنتج: يدعم product_name أو name
        prod_name = r.get("product_name") or r.get("name") or ""

        ws.append([
            idx,
            r.get("product_id") or "",
            r.get("unit_id") or "",
            r.get("code") or "",
            prod_name,
            r.get("unit_name") or "",
            float(qty),
            float(price),
            float(total),
        ])

    # formats
    for rr in range(2, ws.max_row + 1):
        ws.cell(rr, 1).alignment = Alignment(horizontal="center", vertical="center")  # #
        for col in [2, 3, 4, 6]:
            ws.cell(rr, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(rr, 5).alignment = Alignment(horizontal="right", vertical="center")  # المنتج

        ws.cell(rr, 7).number_format = '#,##0.00'
        ws.cell(rr, 8).number_format = '#,##0.00'
        ws.cell(rr, 9).number_format = '#,##0.00'
        for col in [7, 8, 9]:
            ws.cell(rr, col).alignment = Alignment(horizontal="center", vertical="center")

    # ✅ إخفاء الأعمدة التقنية (product_id, unit_id) لو تحب
    ws.column_dimensions["B"].hidden = True  # product_id
    ws.column_dimensions["C"].hidden = True  # unit_id

    # Footer total
    last_data_row = ws.max_row
    footer_row = last_data_row + 1
    ws.cell(footer_row, 8).value = "الإجمالي"
    ws.cell(footer_row, 8).font = Font(bold=True)
    ws.cell(footer_row, 9).value = f"=SUM(I2:I{last_data_row})"
    ws.cell(footer_row, 9).number_format = '#,##0.00'
    ws.cell(footer_row, 9).font = Font(bold=True)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"sales_period_{period_id}.xlsx"
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp






from io import BytesIO
import openpyxl
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.admin.views.decorators import staff_member_required

@staff_member_required
@require_POST
def api_sales_import_xlsx(request):
    f = request.FILES.get("file")
    if not f:
        return JsonResponse({"ok": False, "error": "file مطلوب"}, status=400)

    try:
        wb = openpyxl.load_workbook(filename=BytesIO(f.read()), data_only=True)
    except Exception:
        return JsonResponse({"ok": False, "error": "ملف Excel غير صالح"}, status=400)

    ws = wb.active

    header = [str(ws.cell(1, c).value or "").strip() for c in range(1, 9)]
    expected = ["product_id","unit_id","الكود","المنتج","الوحدة","الكمية","سعر الوحدة","الإجمالي"]
    if header[:8] != expected:
        return JsonResponse({
            "ok": False,
            "error": "تنسيق الملف غير مطابق. استخدم ملف التصدير كقالب (نفس الأعمدة والترتيب)."
        }, status=400)

    rows = []
    count = 0

    for r in range(2, ws.max_row + 1):
        product_id = ws.cell(r, 1).value
        unit_id = ws.cell(r, 2).value
        if product_id in (None, ""):
            continue

        try:
            product_id = int(product_id)
        except Exception:
            continue

        try:
            unit_id = int(unit_id) if unit_id not in (None, "") else None
        except Exception:
            unit_id = None

        qty = _dec(ws.cell(r, 6).value, "0")
        price = _dec(ws.cell(r, 7).value, "0")
        total = qty * price  # ✅ للعرض

        # تجاهل صفوف صفر بالكامل (اختياري)
        if qty == 0 and price == 0:
            continue

        rows.append({
            "line_id": None,
            "product_id": product_id,
            "unit_id": unit_id,
            "code": str(ws.cell(r, 3).value or "").strip(),

            # ✅ خليها product_name لتتطابق مع API عندك (والقالب نعمل له fallback)
            "product_name": str(ws.cell(r, 4).value or "").strip(),

            "unit_name": str(ws.cell(r, 5).value or "").strip(),
            "quantity": str(qty),
            "unit_price": str(price),

            # ✅ لا تتركه 0 (للعرض)
            "line_total": str(total),
        })
        count += 1

    return JsonResponse({"ok": True, "rows": rows, "count": count})
