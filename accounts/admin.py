from django.contrib import admin
from .models import Account, AccountGroup, Mapping, SubMapping, JournalEntry, JournalEntryLine
from django.utils.html import format_html
from django.db.models import Sum, F
from django.http import HttpResponse, HttpResponseRedirect
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
####################################################################################################

@admin.register(Account)
class AccountAdmin(admin.ModelAdmin):
    list_display = ('code', 'name', 'level', 'parent','statement_type')
    list_filter = ('sub_mapping',)
    search_fields = ['code', 'name']
    ordering = ['code']
    def get_mapping(self, obj):
        return obj.sub_mapping.mapping if obj.sub_mapping and obj.sub_mapping.mapping else "-"
    get_mapping.short_description = "التصنيف الرئيسي (Mapping)"

admin.site.register(AccountGroup)
admin.site.register(Mapping)
admin.site.register(SubMapping)



# accounts/admin.py
####################################################################################################
from django.contrib import admin
from .models import AccountDirection

admin.site.register(AccountDirection)

####################################################################################################
# accounts/admin.py

from django.contrib import admin
from .models import FinancialStatementType

admin.site.register(FinancialStatementType)

####################################################################################################
from django.contrib import admin
from django.forms.models import BaseInlineFormSet
from django.http import HttpResponse
from datetime import datetime
import openpyxl
from django.db.models import Q

from .models import OpeningBalance, OpeningBalanceItem
from accounts.models import Account  # تأكد من الاستيراد الصحيح
from django.urls import path
from django.shortcuts import redirect
from django.utils.html import format_html
from django.contrib import messages

class OpeningBalanceItemInlineFormSet(BaseInlineFormSet):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        if not self.instance.pk:
            accounts = Account.objects.filter(
                level__in=[4, 5]
            ).filter(
                Q(code__regex=r'^[123]') |
                Q(sub_mapping__isnull=False)
            ).distinct().order_by('code')
            self.initial = [{'account': acc} for acc in accounts]
            self.extra = len(self.initial)


class OpeningBalanceItemInline(admin.TabularInline):
    model = OpeningBalanceItem
    formset = OpeningBalanceItemInlineFormSet
    extra = 0
    can_delete = False
    fields = ['account', 'debit', 'credit']

    def formfield_for_foreignkey(self, db_field, request=None, **kwargs):
        if db_field.name == 'account':
            accounts = Account.objects.filter(
                level__in=[4, 5]
            ).filter(
                Q(code__regex=r'^[123]') |
                Q(sub_mapping__isnull=False)
            ).distinct().order_by('code')
            kwargs["queryset"] = accounts
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    class Media:
        js = ['js/opening_balance_totals.js']


@admin.register(OpeningBalance)
class OpeningBalanceAdmin(admin.ModelAdmin):
    inlines = [OpeningBalanceItemInline]
    list_display = ['id', 'created_at', 'posted']
    list_filter = ['posted']
    readonly_fields = ['posted', 'created_at']
    actions = ['post_balances', 'unpost_balances', 'export_to_excel']

    def change_view(self, request, object_id, form_url='', extra_context=None):
        obj = self.get_object(request, object_id)
        extra_context = extra_context or {}
        if obj:
            extra_context['show_post'] = not obj.posted
            extra_context['show_unpost'] = obj.posted
        return super().change_view(request, object_id, form_url, extra_context=extra_context)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('<int:pk>/post/', self.admin_site.admin_view(self.process_post), name='openingbalance-post'),
            path('<int:pk>/unpost/', self.admin_site.admin_view(self.process_unpost), name='openingbalance-unpost'),
        ]
        return custom_urls + urls

    def process_post(self, request, pk):
        obj = self.get_object(request, pk)
        if not obj.posted:
            obj.post()
            self.message_user(request, f"✅ تم ترحيل الأرصدة الافتتاحية رقم {obj.id}.")
        return redirect(f'../../{pk}/change/')

    def process_unpost(self, request, pk):
        obj = self.get_object(request, pk)
        if obj.posted:
            obj.unpost()
            self.message_user(request, f"❌ تم إلغاء ترحيل الأرصدة الافتتاحية رقم {obj.id}.")
        return redirect(f'../../{pk}/change/')



    def export_to_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Opening Balances"
        ws.append(['رقم', 'التاريخ', 'تم الترحيل', 'الحساب', 'مدين', 'دائن'])

        for balance in queryset:
            for item in balance.items.all():
                ws.append([
                    balance.id,
                    balance.created_at.strftime('%Y-%m-%d'),
                    "نعم" if balance.posted else "لا",
                    f"{item.account.code} - {item.account.name}",
                    float(item.debit),
                    float(item.credit),
                ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f"opening_balances_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

    export_to_excel.short_description = "📥 تصدير الأرصدة الافتتاحية إلى Excel"

    def post_balances(self, request, queryset):
        posted_count = 0
        for balance in queryset:
            if not balance.posted:
                balance.post()
                posted_count += 1
        self.message_user(request, f"✅ تم ترحيل {posted_count} رصيد/أرصدة بنجاح.")
    post_balances.short_description = "📌 ترحيل الأرصدة الافتتاحية"

    def unpost_balances(self, request, queryset):
        unposted_count = 0
        for balance in queryset:
            if balance.posted:
                balance.unpost()
                unposted_count += 1
        self.message_user(request, f"❌ تم إلغاء ترحيل {unposted_count} رصيد/أرصدة.")
    unpost_balances.short_description = "🚫 إلغاء ترحيل الأرصدة"


###################################################################################################


class JournalEntryLineInline(admin.TabularInline):
    model = JournalEntryLine
    extra = 1
    autocomplete_fields = ['account']
    fields = ['account', 'debit', 'credit', 'description']
    min_num = 2
    verbose_name = "تفصيل القيد"
    verbose_name_plural = "تفاصيل القيد"

    def has_add_permission(self, request, obj=None):
        return not (obj and getattr(obj, 'is_auto', False))

    def has_change_permission(self, request, obj=None):
        return not (obj and getattr(obj, 'is_auto', False))

    def has_delete_permission(self, request, obj=None):
        return not (obj and getattr(obj, 'is_auto', False))

    def get_readonly_fields(self, request, obj=None):
        if obj and getattr(obj, 'is_auto', False):
            return self.fields
        return self.readonly_fields

@admin.register(JournalEntry)
class JournalEntryAdmin(admin.ModelAdmin):

    list_display = ['number', 'date', 'notes', 'colored_debit_total', 'colored_credit_total', 'colored_is_balanced', 'source_type']    
    search_fields = ['number', 'notes']
    inlines = [JournalEntryLineInline]
    date_hierarchy = 'date'
    readonly_fields = ['number']
    exclude = ['reference', 'content_type', 'object_id', 'production_order']


    def colored_debit_total(self, obj):
        total = sum(line.debit for line in obj.lines.all())
        return format_html('<span style="color: green;">💵 {} </span>', "{:,.2f}".format(total))
    colored_debit_total.short_description = "إجمالي المدين"

    def colored_credit_total(self, obj):
        total = sum(line.credit for line in obj.lines.all())
        return format_html('<span style="color: red;">💳 {} </span>', "{:,.2f}".format(total))
    colored_credit_total.short_description = "إجمالي الدائن"

    def colored_is_balanced(self, obj):
        debit = sum(line.debit for line in obj.lines.all())
        credit = sum(line.credit for line in obj.lines.all())
        if debit == credit:
            return format_html('<span style="color: green;">✅ متوازن</span>')
        else:
            return format_html('<span style="color: red;">❌ غير متوازن</span>')
    colored_is_balanced.short_description = "متوازن؟"
    
    def source_type(self, obj):
        if not obj.content_type:
            return format_html('<span style="color: gray;">📝 يدوي</span>')

        model = obj.content_type.model_class()
        model_name = model.__name__.lower()

        mapping = {
            'salesinvoice': '🧾 فاتورة مبيعات',
            'salesreturn': '↩️ مردود مبيعات',
            'purchaseinvoice': '🧾 فاتورة مشتريات',
            'purchasereturn': '↩️ مردود مشتريات',
            'openingbalance': '💰 رصيد افتتاحي',
            'supplierpayment': '💳 دفعة مورد',
            'customerpayment': '💵 دفعة عميل',
            'productionorder': '🏭 أمر إنتاج',
            'stocktransaction': '🚚 حركة مخزون',
            'treasuryvoucher': '💵 حركة صندوق',
        }

        label = mapping.get(model_name, f'🔍 {model.__name__}')
        return format_html('<span style="color: #007bff;">{}</span>', label)

    source_type.short_description = "منشأ من"

    def has_change_permission(self, request, obj=None):
        if obj and obj.is_auto:
            return False
        return super().has_change_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        if obj and obj.is_auto:
            return False
        return super().has_delete_permission(request, obj)
    

####################################################################################################

# accounts/admin.py

from django.db import models
from django.contrib import admin
from django.http import HttpResponseRedirect

class AccountsReportsFakeModel(models.Model):
    class Meta:
        managed = False  # لا يُنشئ جدول فعلي
        app_label = 'accounts'
        verbose_name = "📘 تقارير الحسابات"
        verbose_name_plural = "📘 تقارير الحسابات"

    def __str__(self):
        return "📘 تقارير الحسابات"

@admin.register(AccountsReportsFakeModel)
class AccountsReportsFakeModelAdmin(admin.ModelAdmin):
    def changelist_view(self, request, extra_context=None):
        return HttpResponseRedirect('/accounts/reports/')
